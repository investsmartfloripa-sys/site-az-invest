"""Build do JSON do Painel Visao Geral - bloco CNI.

Indices:
- ICEI - Indice de Confianca do Empresario Industrial (mensal desde 1999)
- INEC - Indice Nacional de Expectativa do Consumidor (mensal)

Fonte: Portal da Industria (CNI) - XLSX publicos.
URLs tipicas:
- https://www.portaldaindustria.com.br/estatisticas/icei-indice-de-confianca-do-empresario-industrial/
- https://www.portaldaindustria.com.br/estatisticas/inec-indice-nacional-de-expectativa-do-consumidor/

Scraping defensivo.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_cni.json"
UA = {"User-Agent": "az-invest-visao-geral-cni/0.1"}

PAGES = {
    "icei": "https://www.portaldaindustria.com.br/estatisticas/icei-indice-de-confianca-do-empresario-industrial/",
    "inec": "https://www.portaldaindustria.com.br/estatisticas/inec-indice-nacional-de-expectativa-do-consumidor/",
}

INPUTS = {"icei": "1999-04", "inec": "1999-09"}


def _get(url: str, *, timeout: int = 60, retries: int = 2) -> requests.Response | None:
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception:
            time.sleep(3)
    return None


def localizar_xlsx(page_url: str) -> str | None:
    r = _get(page_url)
    if r is None:
        return None
    urls = re.findall(r'href=["\']([^"\']+\.xlsx?)["\']', r.text)
    if not urls:
        return None
    u = urls[0]
    if u.startswith("//"):
        u = "https:" + u
    elif u.startswith("/"):
        u = "https://www.portaldaindustria.com.br" + u
    return u


MES_PT = {"JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
          "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12}


def parse_data(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.match(r"^([A-Za-z]{3})[/\-\s](\d{2,4})$", s)
    if m:
        mes = MES_PT.get(m.group(1).upper())
        if mes:
            ano = int(m.group(2))
            ano = ano + 2000 if ano < 100 else ano
            return f"{ano:04d}-{mes:02d}"
    return None


def parse_xlsx(content: bytes) -> dict[str, float]:
    from openpyxl import load_workbook
    out: dict[str, float] = {}
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return out
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        # acha linha com data na col 0 e numero na col 1+
        for row in rows:
            if not row or len(row) < 2:
                continue
            mes = parse_data(row[0])
            if not mes:
                continue
            # primeira coluna numerica
            for c in row[1:]:
                if isinstance(c, (int, float)) and not isinstance(c, bool):
                    out[mes] = float(c)
                    break
        if out:
            break
    return out


def serie_lista(d: dict[str, float]) -> list[dict]:
    return [{"mes": m, "valor": d[m]} for m in sorted(d.keys())]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "visao_geral_cni.json"

    print("== CNI - ICEI + INEC ==")
    series: dict[str, dict[str, float]] = {}
    for slug, page in PAGES.items():
        print(f"  -> {slug.upper()}")
        url = localizar_xlsx(page)
        if not url:
            print("    XLSX nao localizado")
            series[slug] = {}
            continue
        print(f"    XLSX: {url}")
        r = _get(url)
        series[slug] = parse_xlsx(r.content) if r is not None else {}
        print(f"    {len(series[slug])} obs")

    if not any(series.values()):
        sys.path.insert(0, str(HERE))
        from shared.blob_download import download_json
        prev = download_json(BLOB_PATH)
        if prev:
            prev["freshness_status"] = "stale"
            prev["gerado_em"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            out_file.write_text(json.dumps(prev, indent=2, ensure_ascii=False), encoding="utf-8")
            return
        payload = {
            "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "freshness_status": "missing",
            "icei": [],
            "inec": [],
        }
        out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return

    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "freshness_status": "fresh",
        "icei": serie_lista(series["icei"]),
        "inec": serie_lista(series["inec"]),
        "inputs": INPUTS,
        "min_start_date": max(INPUTS.values()),
        "metadata": {
            "fonte": "CNI / Portal da Industria. ICEI (industrial) mensal desde 1999; INEC (consumidor) mensal.",
            "nota": "50 = neutro para INEC; 50 e linha de corte tambem para ICEI (>50 otimismo).",
        },
    }
    out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")

    if args.upload:
        sys.path.insert(0, str(HERE))
        from shared.blob_upload import maybe_upload_json
        try:
            maybe_upload_json(out_file, BLOB_PATH)
        except Exception as e:
            print(f"[upload] FALHOU: {e}", file=sys.stderr)
            if not args.soft_fail:
                sys.exit(1)


if __name__ == "__main__":
    main()
