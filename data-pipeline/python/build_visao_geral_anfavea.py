"""Build do Painel Visao Geral - bloco ANFAVEA (producao + vendas + exportacao).

Layout WIDE da ANFAVEA: cada XLSX <siteautoveiculos{ANO}.xlsx> contem 12 meses do ano.
Loop sobre anos 2019..corrente para gerar serie historica completa.

Inputs:
  https://anfavea.com.br/docs/siteautoveiculos{ANO}.xlsx (2019..2026 confirmados HTTP 200)

Outputs:
  data/visao_geral_anfavea.json com serie mensal (producao/vendas/exportacao_unidades,
  YoY pct, indice base 2019).
"""
from __future__ import annotations
import argparse, json, re, sys, time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import requests

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_anfavea.json"
UA = {"User-Agent": "Mozilla/5.0 az-invest-anfavea"}
ANFAVEA_URL_TPL = "https://anfavea.com.br/docs/siteautoveiculos{ano}.xlsx"
INPUTS = {"anfavea_veiculos": "2019-01"}

# Map aba -> tipo. Layout WIDE: linha header com Jan..Dez, linha "Total" abaixo.
SHEET_TO_KIND = {
    "VI. PRODUCAO": "producao",
    "VI. PRODUCÃO": "producao",
    "VI. PRODUÇÃO": "producao",
    "I. EMPLACAMENTO": "vendas",
    "V. EXPORTACAO VOLUME": "exportacao",
    "V. EXPORTAÇÃO VOLUME": "exportacao",
}
MESES_PT = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def _get(url, *, timeout=60, retries=3, sleep=3.0):
    last = None
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception as e:
            last = e
            time.sleep(sleep)
    raise RuntimeError(f"falha {url}: {last}")


def parse_xlsx_wide(content: bytes, ano: int) -> list[dict]:
    """Layout WIDE: para cada aba (Produção/Emplacamento/Exportação) acha a linha header
    com Jan..Dez e depois a linha 'Total' nas linhas seguintes, captura os 12 numeros."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    by_mes: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        key = sheet_name.upper().strip()
        kind = None
        # match flexivel
        for k, v in SHEET_TO_KIND.items():
            if k in key or key in k:
                kind = v
                break
        if kind is None:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # achar linha com Jan..Dez nas primeiras 15 linhas
        meses_row_idx = None
        for i, row in enumerate(rows[:15]):
            ups = [str(c).strip().upper()[:3] if c else "" for c in row]
            if "JAN" in ups and "DEZ" in ups:
                meses_row_idx = i
                break
        if meses_row_idx is None:
            continue

        col_to_mes: dict[int, int] = {}
        for j, c in enumerate(rows[meses_row_idx]):
            if c is None:
                continue
            tag = str(c).strip().upper()[:3]
            if tag in MESES_PT:
                col_to_mes[j] = MESES_PT.index(tag) + 1

        # achar linha "Total" abaixo
        total_row = None
        for k in range(meses_row_idx + 1, min(meses_row_idx + 8, len(rows))):
            row_k = rows[k]
            for c in row_k[:4]:
                if c and str(c).strip().upper() == "TOTAL":
                    total_row = row_k
                    break
            if total_row:
                break
        if total_row is None:
            continue

        for col, mes in col_to_mes.items():
            try:
                v = total_row[col]
                if v is None:
                    continue
                v = float(v)
            except (TypeError, ValueError, IndexError):
                continue
            if v <= 0:
                continue
            mes_iso = f"{ano:04d}-{mes:02d}"
            entry = by_mes.setdefault(mes_iso, {"mes": mes_iso})
            entry[f"{kind}_unidades"] = v

    return list(by_mes.values())


def calcular(serie: list[dict]) -> list[dict]:
    """Calcula base_2019 + YoY mes a mes + indice 2019 + producao/vendas."""
    base = {}
    for it in serie:
        if it["mes"].startswith("2019-"):
            for k in ("producao_unidades", "vendas_unidades", "exportacao_unidades"):
                if it.get(k):
                    base.setdefault(k, []).append(it[k])
    base = {k: (sum(v) / len(v) if v else None) for k, v in base.items()}

    by_mes = {it["mes"]: it for it in serie}
    for it in serie:
        a, m = it["mes"].split("-")
        prev = by_mes.get(f"{int(a) - 1:04d}-{m}")
        for k in ("producao_unidades", "vendas_unidades", "exportacao_unidades"):
            short = k.replace("_unidades", "")
            cur = it.get(k)
            pv = prev.get(k) if prev else None
            it[f"{short}_var_yoy_pct"] = round((cur / pv - 1) * 100, 2) if (cur and pv and pv > 0) else None
            bs = base.get(k)
            it[f"{short}_indice_2019"] = round(cur / bs * 100, 2) if (cur and bs and bs > 0) else None
        p = it.get("producao_unidades")
        v = it.get("vendas_unidades")
        it["producao_sobre_vendas"] = round(p / v, 3) if (p and v) else None
    return serie


def main():
    ap = argparse.ArgumentParser(description="Build Painel Visao Geral - ANFAVEA")
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ano_atual = datetime.now(timezone.utc).year
    anos = list(range(2019, ano_atual + 1))

    serie_total: dict[str, dict] = {}
    n_ok = 0
    for ano in anos:
        url = ANFAVEA_URL_TPL.format(ano=ano)
        try:
            r = _get(url, retries=2)
            for item in parse_xlsx_wide(r.content, ano):
                serie_total[item["mes"]] = {**serie_total.get(item["mes"], {}), **item}
            n_ok += 1
            print(f"  ano {ano}: ok ({len([m for m in serie_total if m.startswith(str(ano))])} meses)")
        except Exception as e:
            print(f"  ano {ano}: FALHOU ({e})", file=sys.stderr)

    if not serie_total:
        msg = "Nenhum ano ANFAVEA baixado"
        if args.soft_fail:
            print(f"SOFT-FAIL: {msg}", file=sys.stderr)
            sys.exit(0)
        raise RuntimeError(msg)

    serie = [serie_total[m] for m in sorted(serie_total)]
    serie = calcular(serie)

    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(),
        "freshness_status": "fresh" if n_ok >= len(anos) - 1 else "stale",
        "mes_recente": serie[-1]["mes"] if serie else None,
        "serie": serie,
        "inputs": INPUTS,
        "min_start_date": "2019-01",
        "metadata": {
            "fonte": "ANFAVEA - siteautoveiculos{ANO}.xlsx",
            "nota": f"Layout WIDE. {n_ok}/{len(anos)} anos baixados. Base 2019 para indice. YoY contra mesmo mes ano anterior.",
        },
    }

    out_path = out_dir / "visao_geral_anfavea.json"
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"  -> {out_path} ({len(serie)} obs)")

    if args.upload:
        try:
            sys.path.insert(0, str(HERE / "shared"))
            from blob_upload import maybe_upload_json
            maybe_upload_json(out_path, BLOB_PATH)
        except Exception as e:
            print(f"upload skip: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
