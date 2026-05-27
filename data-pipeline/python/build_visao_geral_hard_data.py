"""Build do JSON do Painel Visao Geral - bloco Hard Data fisico (Onda 2).

Agrega em um unico JSON varios indicadores antecedentes fisicos de alta
frequencia, todos coletados via scraping defensivo (fragil por natureza):
- ABCR - fluxo de veiculos pedagiados (leves/pesados)
- ABPO - expedicao de papelao ondulado (Empapel)
- SNIC - vendas de cimento
- Instituto Aco Brasil - producao + vendas internas de aco bruto
- FENABRAVE - emplacamentos (cross-check do ANFAVEA vendas)

Cada fonte e tentada independentemente; falha pontual nao derruba o resto.
Padrao: 1 dict por fonte com {serie: [{mes, valor}], freshness_status,
last_url_used}. JSON consolidado.

Como o formato exato dos PDFs/XLSX varia, esse pipeline e a versao mais
fragil do projeto. Recomenda-se sempre rodar com --soft-fail e ter
revisao manual periodica.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_hard_data.json"
UA = {"User-Agent": "az-invest-visao-geral-hardd/0.1"}

PAGES = {
    "abcr": "https://www.melhoresrodovias.org.br/indice-abcr/",
    "abpo": "https://www.empapel.org.br/publicacoes/boletim-estatistico/",
    "snic": "http://snic.org.br/",
    "aco": "https://acobrasil.org.br/site/estatistica-mensal/",
    "fenabrave": "https://www.fenabrave.org.br/",
}

INPUTS = {
    "abcr": "1999-01",
    "abpo": "2000-01",
    "snic": "2000-01",
    "aco": "2000-01",
    "fenabrave": "2000-01",
}


def _get(url: str, *, timeout: int = 60, retries: int = 2) -> requests.Response | None:
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception:
            time.sleep(3)
    return None


def localizar_anexo(page_url: str, extensoes: tuple[str, ...]) -> str | None:
    r = _get(page_url)
    if r is None:
        return None
    pattern = r'href=["\']([^"\']+(?:' + "|".join(re.escape(e) for e in extensoes) + r'))["\']'
    urls = re.findall(pattern, r.text, re.IGNORECASE)
    if not urls:
        return None
    base = "/".join(page_url.split("/")[:3])
    u = urls[0]
    if u.startswith("//"):
        u = "https:" + u
    elif u.startswith("/"):
        u = base + u
    return u


MES_PT = {"JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
          "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12}


def parse_data(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.match(r"^([A-Za-z]{3})[/\-\s](\d{2,4})$", s)
    if m:
        mes = MES_PT.get(m.group(1).upper())
        if mes:
            ano = int(m.group(2))
            ano = ano + 2000 if ano < 100 else ano
            return f"{ano:04d}-{mes:02d}"
    return None


def parse_xlsx_simples(content: bytes) -> dict[str, float]:
    """Pega coluna 0 = data, primeira coluna numerica = valor."""
    from openpyxl import load_workbook

    out: dict[str, float] = {}
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return out
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        for row in rows:
            if not row or len(row) < 2:
                continue
            mes = parse_data(row[0])
            if not mes:
                continue
            for c in row[1:]:
                if isinstance(c, (int, float)) and not isinstance(c, bool):
                    out[mes] = float(c)
                    break
        if out:
            break
    return out


def calcular_yoy(d: dict[str, float]) -> list[dict]:
    meses = sorted(d.keys())
    out = []
    for i, m in enumerate(meses):
        v = d[m]
        prev = d.get(meses[i - 12]) if i >= 12 else None
        yoy = round((v / prev - 1) * 100, 2) if (prev and prev > 0) else None
        out.append({"mes": m, "valor": v, "var_yoy_pct": yoy})
    return out


def fetch_xlsx_pipeline(fonte: str, palavras_filtro: list[str] | None = None) -> dict[str, float]:
    """Pega XLSX da pagina da fonte e parseia."""
    page = PAGES.get(fonte)
    if not page:
        return {}
    url = localizar_anexo(page, (".xlsx", ".xls"))
    if not url:
        print(f"  {fonte}: nao localizou XLSX")
        return {}
    print(f"  {fonte}: XLSX {url}")
    r = _get(url)
    if r is None:
        return {}
    return parse_xlsx_simples(r.content)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "visao_geral_hard_data.json"

    print("== Hard Data fisico (ABCR / ABPO / SNIC / Aco / FENABRAVE) ==")
    sys.path.insert(0, str(HERE))
    from shared.blob_download import download_json

    prev = download_json(BLOB_PATH) or {}

    resultado: dict[str, dict] = {}
    for fonte in PAGES:
        print(f"  -> {fonte.upper()}")
        try:
            d = fetch_xlsx_pipeline(fonte)
        except Exception as e:
            print(f"    erro: {e}", file=sys.stderr)
            d = {}

        if d:
            resultado[fonte] = {
                "serie": calcular_yoy(d),
                "freshness_status": "fresh",
            }
            print(f"    {len(d)} obs")
        else:
            anterior = prev.get(fonte)
            if anterior and anterior.get("serie"):
                resultado[fonte] = {**anterior, "freshness_status": "stale"}
                print(f"    preservado anterior ({len(anterior['serie'])} obs, stale)")
            else:
                resultado[fonte] = {"serie": [], "freshness_status": "missing"}
                print("    sem dado disponivel")

    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "freshness_status": "fresh" if any(r.get("freshness_status") == "fresh" for r in resultado.values()) else "stale",
        **resultado,
        "inputs": INPUTS,
        "min_start_date": max(INPUTS.values()),
        "metadata": {
            "fontes": {
                "abcr": "ABCR - Indice de fluxo de veiculos em rodovias pedagiadas (leves + pesados separados).",
                "abpo": "Empapel/ABPO - Expedicao de papelao ondulado (toneladas).",
                "snic": "SNIC - Vendas de cimento.",
                "aco": "Instituto Aco Brasil - Producao bruta + vendas internas.",
                "fenabrave": "FENABRAVE - Emplacamentos.",
            },
            "nota": "Antecedentes industriais classicos. Parser defensivo - pode falhar quando layouts mudam. JSON anterior preservado como 'stale'.",
        },
    }
    out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")

    if args.upload:
        sys.path.insert(0, str(HERE))
        from shared.blob_upload import maybe_upload_json
        try:
            maybe_upload_json(out_file, BLOB_PATH)
        except Exception as e:
            print(f"[upload] FALHOU: {e}", file=sys.stderr)
            if not args.soft_fail:
                sys.exit(1)


if __name__ == "__main__":
    main()
