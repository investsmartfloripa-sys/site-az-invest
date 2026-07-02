"""Build do JSON do Painel Visao Geral - bloco Antecedentes FGV-IBRE.

Fontes (todas via portal IBRE / FGVDados publico):
- IACE - Indicador Antecedente Composto da Economia (antecedente)
- ICCE - Indicador Coincidente Composto da Economia (coincidente)
- IAEmp - Indicador Antecedente de Emprego
- IIE-Br - Indicador de Incerteza da Economia (mensal + diario)

URLs publicas tipicas:
- https://portalibre.fgv.br/iace-e-icce
- https://portalibre.fgv.br/en/leading-indicator-employment
- https://portalibre.fgv.br/en/iie-br
- https://portalibre.fgv.br/en/daily-measurement-iie-br

Scraping defensivo: procura links de XLSX/PDF nos portais; se nao achar,
preserva JSON anterior do Blob com freshness_status='stale'.

NOTA: FGV-IBRE nao tem API publica. Algumas series so estao disponiveis
via FGVDados (assinatura). Esse pipeline tenta endpoints publicos conhecidos
mas e estruturalmente fragil - parser defensivo + soft-fail.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests

try:
    # portalibre.fgv.br BLOQUEIA o fingerprint TLS do Python (SSLV3 alert
    # handshake failure) mas aceita browsers — curl_cffi impersona o Chrome.
    from curl_cffi import requests as curl_requests
except ImportError:  # ambiente sem curl_cffi: cai no requests (vai falhar na FGV)
    curl_requests = None

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_fgv_antecedentes.json"
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36", "Accept": "*/*"}

PAGES = {
    "iace_icce": "https://portalibre.fgv.br/iace-e-icce",
    "iaemp": "https://portalibre.fgv.br/en/leading-indicator-employment",
    "iie_br": "https://portalibre.fgv.br/en/iie-br",
}

INPUTS = {
    "iace": "1996-01",
    "icce": "1996-01",
    "iaemp": "2008-06",
    "iie_br": "2000-01",
}


def _get(url: str, *, timeout: int = 60, retries: int = 2, sleep: float = 4.0):
    """GET com impersonação de browser (curl_cffi) e fallback pro requests."""
    for i in range(retries):
        try:
            if curl_requests is not None:
                r = curl_requests.get(url, timeout=timeout, impersonate="chrome", headers=UA)
            else:
                r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception as e:
            print(f"  retry {i + 1}/{retries} {url}: {e}", file=sys.stderr)
            time.sleep(sleep)
    return None


def localizar_xlsx_na_pagina(page_url: str, palavras: list[str]) -> str | None:
    """Acha link de XLSX na pagina que tenha pelo menos uma das palavras-chave."""
    r = _get(page_url)
    if r is None:
        return None
    urls = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', r.text)
    if not urls:
        return None
    for u in urls:
        ul = u.lower()
        if any(p in ul for p in palavras):
            if u.startswith("//"):
                u = "https:" + u
            elif u.startswith("/"):
                u = "https://portalibre.fgv.br" + u
            return u
    # se nao casou palavras, retorna o primeiro
    u = urls[0]
    if u.startswith("//"):
        u = "https:" + u
    elif u.startswith("/"):
        u = "https://portalibre.fgv.br" + u
    return u


MES_PT = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}


def parse_data_celula(v) -> str | None:
    """Tenta interpretar varias formas: datetime, '2024-04', 'abr/24', 04/2024, '2024-04-01'."""
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    if not s:
        return None
    # ISO: 2024-04 ou 2024-04-01
    m = re.match(r"^(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    # abr/24, jan/2024, jul-2024
    m = re.match(r"^([A-Za-z]{3,9})[/\-\s](\d{2,4})$", s)
    if m:
        mes_str = m.group(1).upper().replace("Ç", "C")
        mes = MES_PT.get(mes_str)
        if mes:
            ano = int(m.group(2))
            ano = ano + 2000 if ano < 100 else ano
            return f"{ano:04d}-{mes:02d}"
    # mm/yyyy
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"
    return None


def parse_xlsx_generico(content: bytes, indicadores_esperados: list[str]) -> dict[str, dict[str, float | None]]:
    """Parser defensivo de XLSX FGV.

    Estrategia: percorre todas as abas, identifica uma coluna de data e
    colunas numericas com nomes que casem com `indicadores_esperados`.
    Retorna {indicador: {mes_iso: valor}}.
    """
    from openpyxl import load_workbook

    out: dict[str, dict[str, float | None]] = {ind: {} for ind in indicadores_esperados}
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        print(f"  XLSX abrir: {e}", file=sys.stderr)
        return out

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Procura linha de header
        header_idx = None
        for i, row in enumerate(rows[:25]):
            row_str = [str(c).upper() if c is not None else "" for c in row]
            joined = " ".join(row_str)
            if any(ind.upper() in joined for ind in indicadores_esperados):
                header_idx = i
                break
        if header_idx is None:
            continue

        header = [str(c).strip().upper() if c is not None else "" for c in rows[header_idx]]
        # detectar coluna de data (geralmente coluna 0 ou contém 'DATA'/'MES'/'PERIODO')
        col_data = next(
            (i for i, c in enumerate(header) if "DATA" in c or "MES" in c or "PERIO" in c or i == 0),
            0,
        )

        # mapear indicadores -> colunas
        mapa: dict[str, int] = {}
        for ind in indicadores_esperados:
            for i, c in enumerate(header):
                if ind.upper() in c and i != col_data:
                    mapa[ind] = i
                    break

        if not mapa:
            continue

        for row in rows[header_idx + 1 :]:
            if len(row) <= col_data:
                continue
            mes_iso = parse_data_celula(row[col_data])
            if not mes_iso:
                continue
            for ind, col in mapa.items():
                if col >= len(row):
                    continue
                v = row[col]
                if v is None or isinstance(v, str) and not v.strip():
                    continue
                try:
                    out[ind][mes_iso] = float(v)
                except (TypeError, ValueError):
                    continue
    return out


def variacao_yoy(serie: dict[str, float | None]) -> dict[str, float | None]:
    meses = sorted(serie.keys())
    out: dict[str, float | None] = {}
    for i, m in enumerate(meses):
        if i < 12:
            out[m] = None
            continue
        cur = serie.get(m)
        prev = serie.get(meses[i - 12])
        if cur is None or prev is None or prev == 0:
            out[m] = None
            continue
        out[m] = round((cur / prev - 1) * 100, 2)
    return out


def serie_lista(d: dict[str, float | None]) -> list[dict]:
    return [{"mes": m, "valor": d[m]} for m in sorted(d.keys())]


def main() -> None:
    ap = argparse.ArgumentParser(description="Build do JSON Painel Visao Geral - FGV antecedentes")
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "visao_geral_fgv_antecedentes.json"

    print("== FGV-IBRE - antecedentes (IACE/ICCE/IAEmp/IIE-Br) ==")
    series: dict[str, dict[str, float | None]] = {}

    # IACE + ICCE
    print("  -> IACE / ICCE")
    url_iace = localizar_xlsx_na_pagina(PAGES["iace_icce"], ["iace", "icce", "indicador"])
    if url_iace:
        print(f"    XLSX: {url_iace}")
        r = _get(url_iace)
        if r is not None:
            parsed = parse_xlsx_generico(r.content, ["IACE", "ICCE"])
            series.update(parsed)
    else:
        print("    nao localizado")

    # IAEmp
    print("  -> IAEmp")
    url_iaemp = localizar_xlsx_na_pagina(PAGES["iaemp"], ["iaemp", "antecedente", "emprego", "leading"])
    if url_iaemp:
        print(f"    XLSX: {url_iaemp}")
        r = _get(url_iaemp)
        if r is not None:
            parsed = parse_xlsx_generico(r.content, ["IAEMP"])
            series.update(parsed)
    else:
        print("    nao localizado")

    # IIE-Br
    print("  -> IIE-Br")
    url_iiebr = localizar_xlsx_na_pagina(PAGES["iie_br"], ["iie", "incerteza", "uncertainty"])
    if url_iiebr:
        print(f"    XLSX: {url_iiebr}")
        r = _get(url_iiebr)
        if r is not None:
            parsed = parse_xlsx_generico(r.content, ["IIE-BR", "IIE_BR", "IIEBR", "IIE BR"])
            series.update(parsed)
    else:
        print("    nao localizado")

    # se nada veio, soft-fail — SEM return precoce: o marcador stale/missing
    # também precisa SUBIR pro Blob (bug antigo: o return pulava o upload e a
    # fonte ficava 404 eterna no monitoramento).
    any_serie = any(series.values())
    if not any_serie:
        print("  Nenhuma serie obtida — soft-fail / preservar Blob", file=sys.stderr)
        sys.path.insert(0, str(HERE))
        from shared.blob_download import download_json
        prev = download_json(BLOB_PATH)
        if prev:
            payload = dict(prev)
            payload["freshness_status"] = "stale"
            payload["gerado_em"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        else:
            payload = {
                "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "freshness_status": "missing",
                "iace": [],
                "icce": [],
                "iaemp": [],
                "iie_br": [],
                "metadata": {"fonte": "FGV-IBRE", "nota": "Scraper nao localizou XLSX nas paginas conhecidas. Manter Onda 2 sob revisao manual."},
            }
        out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"JSON {out_file} (soft-fail, freshness={payload['freshness_status']})")
    else:
        # normaliza chaves
        def pegar(*keys: str) -> dict[str, float | None]:
            for k in keys:
                if k in series and series[k]:
                    return series[k]
            return {}

        iace = pegar("IACE")
        icce = pegar("ICCE")
        iaemp = pegar("IAEMP")
        iie_br = pegar("IIE-BR", "IIE_BR", "IIEBR", "IIE BR")

        payload = {
            "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "freshness_status": "fresh",
            "iace": {"serie": serie_lista(iace), "var_yoy": serie_lista(variacao_yoy(iace))},
            "icce": {"serie": serie_lista(icce), "var_yoy": serie_lista(variacao_yoy(icce))},
            "iaemp": {"serie": serie_lista(iaemp), "var_yoy": serie_lista(variacao_yoy(iaemp))},
            "iie_br": {"serie": serie_lista(iie_br), "var_yoy": serie_lista(variacao_yoy(iie_br))},
            "inputs": INPUTS,
            "min_start_date": max(INPUTS.values()),
            "metadata": {
                "fonte": "FGV-IBRE - portal pubblico. IACE/ICCE (Conference Board), IAEmp, IIE-Br.",
                "nota": "Scraping de XLSX (TLS impersonado via curl_cffi — o portal FGV rejeita o TLS padrao do Python). Em caso de falha, JSON anterior preservado como 'stale'.",
            },
        }

        out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"JSON {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")
        print(f"  IACE {len(iace)} obs | ICCE {len(icce)} | IAEmp {len(iaemp)} | IIE-Br {len(iie_br)}")

    if args.upload:
        sys.path.insert(0, str(HERE))
        from shared.blob_upload import maybe_upload_json
        try:
            maybe_upload_json(out_file, BLOB_PATH)
        except Exception as e:
            print(f"[upload] FALHOU: {e}", file=sys.stderr)
            if not args.soft_fail:
                sys.exit(1)


if __name__ == "__main__":
    main()
