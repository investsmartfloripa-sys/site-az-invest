"""Build do JSON do Painel Visão Geral — bloco Consumo de Energia Elétrica (EPE).

Fonte: Resenha Mensal do Mercado de Energia Elétrica da EPE em
https://www.epe.gov.br/pt/publicacoes-dados-abertos/publicacoes/consumo-de-energia-eletrica

A EPE publica XLSX consolidado mensal com consumo por classe (residencial,
industrial, comercial, outros) em GWh, desde 1970.

Calcula:
- Variação a/a por classe
- MM3m
- Índice base 2019=100

O consumo industrial é antecedente FORTÍSSIMO da PIM — quando cai
sustentado, sinal de freio industrial em 1-2 meses.

Ragged-edge tolerante.
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_epe.json"
UA = {"User-Agent": "az-invest-visao-geral-epe/0.1"}

EPE_PAGE = "https://www.epe.gov.br/pt/publicacoes-dados-abertos/publicacoes/consumo-de-energia-eletrica"
# URL fallback (XLSX consolidado publicado pela EPE). Nome muda; mantemos referência.
EPE_FALLBACK_XLSX = "https://www.epe.gov.br/sites-pt/publicacoes-dados-abertos/publicacoes/PublicacoesArquivos/publicacao-153/RESENHA-MENSAL-MERCADO-ENERGIA-ELETRICA.xlsx"

INPUTS = {"epe_consumo": "1970-01"}


def _get(url: str, *, timeout: int = 90, retries: int = 2, sleep: float = 5.0) -> requests.Response:
    last: Exception | None = None
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception as e:
            last = e
            print(f"  retry {i + 1}/{retries}: {e}", file=sys.stderr)
            time.sleep(sleep)
    raise RuntimeError(f"falha após {retries} tentativas: {last}")


def localizar_xlsx() -> str:
    try:
        r = _get(EPE_PAGE, retries=1)
        import re

        urls = re.findall(r'href="([^"]+\.xlsx)"', r.text)
        # priorizar "resenha" ou "consumo"
        for u in urls:
            ul = u.lower()
            if "resenha" in ul or "consumo" in ul or "mercado" in ul:
                if u.startswith("//"):
                    u = "https:" + u
                if u.startswith("/"):
                    u = "https://www.epe.gov.br" + u
                return u
    except Exception as e:
        print(f"  não achou XLSX no portal EPE: {e}", file=sys.stderr)
    return EPE_FALLBACK_XLSX


MES_PT = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}


def parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    print(f"  abas: {wb.sheetnames}")
    # EPE Resenha Mensal costuma ter abas "Consumo Mensal" ou "1.3-Consumo Total" com colunas
    # ANO, MÊS, RESIDENCIAL, INDUSTRIAL, COMERCIAL, OUTROS (em GWh).
    by_mes: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        nome_upper = sheet_name.upper()
        if not ("CONSUMO" in nome_upper or "CLASSE" in nome_upper or "TOTAL" in nome_upper or "BRASIL" in nome_upper):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows[:30]):
            row_str = [str(c).upper() if c else "" for c in row]
            if any("RESIDENC" in c for c in row_str) and any("INDUSTR" in c for c in row_str):
                header_idx = i
                break
        if header_idx is None:
            continue
        header = [str(c).upper() if c else "" for c in rows[header_idx]]

        def col(*keys: str) -> int | None:
            for k in keys:
                for i, c in enumerate(header):
                    if k in c:
                        return i
            return None

        col_ano = col("ANO", "YEAR")
        col_mes = col("MÊS", "MES", "MONTH")
        col_res = col("RESIDENC")
        col_ind = col("INDUSTR")
        col_com = col("COMERC")
        col_out = col("OUTROS", "OTHER")
        col_tot = col("TOTAL")
        if col_ano is None or col_mes is None:
            continue

        for row in rows[header_idx + 1 :]:
            try:
                ano_raw = row[col_ano]
                mes_raw = row[col_mes]
                if ano_raw is None or mes_raw is None:
                    continue
                ano = int(ano_raw)
                if isinstance(mes_raw, str):
                    mes = MES_PT.get(mes_raw.upper().strip())
                    if mes is None:
                        try:
                            mes = int(mes_raw)
                        except ValueError:
                            continue
                else:
                    mes = int(mes_raw)
            except (TypeError, ValueError):
                continue
            mes_iso = f"{ano:04d}-{mes:02d}"
            entry = by_mes.setdefault(mes_iso, {"mes": mes_iso})

            def get(idx: int | None) -> float | None:
                if idx is None:
                    return None
                try:
                    return float(row[idx])
                except (TypeError, ValueError):
                    return None

            entry["residencial_gwh"] = get(col_res) or entry.get("residencial_gwh")
            entry["industrial_gwh"] = get(col_ind) or entry.get("industrial_gwh")
            entry["comercial_gwh"] = get(col_com) or entry.get("comercial_gwh")
            entry["outros_gwh"] = get(col_out) or entry.get("outros_gwh")
            entry["total_gwh"] = get(col_tot) or entry.get("total_gwh")
        if any(by_mes.values()):
            break  # parou na primeira aba relevante

    return [by_mes[m] for m in sorted(by_mes.keys())]


def calcular_variacoes(serie: list[dict]) -> None:
    by_mes = {item["mes"]: item for item in serie}
    keys = ("residencial_gwh", "industrial_gwh", "comercial_gwh", "outros_gwh", "total_gwh")
    bases: dict[str, list[float]] = {k: [] for k in keys}
    for item in serie:
        if item["mes"].startswith("2019-"):
            for k in keys:
                if item.get(k):
                    bases[k].append(item[k])
    base_2019 = {k: (sum(v) / len(v) if v else None) for k, v in bases.items()}

    for item in serie:
        ano, m = item["mes"].split("-")
        prev = by_mes.get(f"{int(ano) - 1:04d}-{m}")
        for k in keys:
            short = k.replace("_gwh", "")
            atual = item.get(k)
            prev_v = prev.get(k) if prev else None
            if atual is not None and prev_v is not None and prev_v > 0:
                item[f"{short}_var_yoy_pct"] = round((atual / prev_v - 1) * 100, 2)
            else:
                item[f"{short}_var_yoy_pct"] = None
            base = base_2019.get(k)
            if atual is not None and base is not None and base > 0:
                item[f"{short}_indice_2019"] = round(atual / base * 100, 2)
            else:
                item[f"{short}_indice_2019"] = None


def main() -> None:
    ap = argparse.ArgumentParser(description="Build do JSON Painel Visão Geral — EPE")
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "visao_geral_epe.json"

    print("== EPE — Resenha Mensal de Energia Elétrica ==")
    try:
        url = localizar_xlsx()
        print(f"  XLSX: {url}")
        r = _get(url)
        serie = parse_xlsx(r.content)
        if not serie:
            raise RuntimeError("XLSX EPE parseado mas série vazia")
    except Exception as e:
        print(f"  FALHA EPE: {e}", file=sys.stderr)
        sys.path.insert(0, str(HERE))
        from shared.blob_download import download_json
        prev = download_json(BLOB_PATH)
        if prev:
            prev["freshness_status"] = "stale"
            prev["gerado_em"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            out_file.write_text(json.dumps(prev, indent=2, ensure_ascii=False), encoding="utf-8")
            print("  preservado JSON anterior (stale)")
            return
        if args.soft_fail:
            payload = {"gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"), "freshness_status": "missing", "serie": []}
            out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
            return
        sys.exit(2)

    calcular_variacoes(serie)

    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "freshness_status": "fresh",
        "mes_recente": serie[-1]["mes"] if serie else None,
        "serie": serie,
        "inputs": INPUTS,
        "min_start_date": min(INPUTS.values()),
        "metadata": {
            "fonte": "EPE — Resenha Mensal do Mercado de Energia Elétrica (XLSX). Consumo em GWh por classe.",
            "nota": "Industrial é antecedente fortíssimo da PIM. Base índice = média 2019.",
        },
    }
    out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")

    if args.upload:
        sys.path.insert(0, str(HERE))
        from shared.blob_upload import maybe_upload_json
        try:
            maybe_upload_json(out_file, BLOB_PATH)
        except Exception as e:
            print(f"[upload] FALHOU: {e}", file=sys.stderr)
            if not args.soft_fail:
                sys.exit(1)


if __name__ == "__main__":
    main()
