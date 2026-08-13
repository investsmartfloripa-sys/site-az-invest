"""Build do JSON de fatores condicionantes da DBGG (decomposição da dívida bruta).

Fonte: XLSX paramétrico do BCB "Tabelas de estatísticas fiscais"
  https://www.bcb.gov.br/content/estatisticas/hist_estatisticasfiscais/{YYYYMM}_Tabelas_de_estatisticas_fiscais.xlsx
  (YYYYMM = mês de PUBLICAÇÃO; safra mais antiga disponível = 201805)

- Tabela 18 = fluxos MENSAIS (~5 meses por safra, R$ milhões | % PIB)
- Tabela 19 = acumulado NO ANO (colunas Dez N-2 / Dez N-1 + meses recentes de N)

Os fatores condicionantes NÃO existem no SGS — só neste XLSX. O parser é por
RÓTULO da coluna A (nunca índice fixo de linha): se o BCB inserir uma linha,
o build aborta em vez de ler juros como emissões.

Decomposição publicada:
  variação DBGG = emissões líquidas + juros nominais + ajuste cambial + outros
  (outros = dívida externa-outros ajustes + reconhecimento de dívidas + privatizações)
  No anual (% PIB) entra também o efeito crescimento do PIB (denominador):
  variação da razão DBGG/PIB (pp) = fatores (% PIB) + efeito PIB (pp)

Validações que ABORTAM o upload:
- soma dos fatores R$ ≈ variação DBGG da própria tabela (tol. R$ 1 mi)
- mensal vs SGS 13761 (DBGG saldos R$ mi): Δ13761 ≈ variação (tol. R$ 5 mi)
- anual vs SGS 13762 (DBGG % PIB): Δdez ≈ variação em pp (tol. 0,05 pp)
  (checks SGS aplicados apenas onde o mês/ano existe na série SGS)

Reconstrução histórica: mensal desde dez/2017 e anual desde 2016, iterando
safras {ano}01/{ano}06/{ano}11 de 2018 a hoje (+ a seed 201805, mais antiga
disponível, que cobre dez/2017) + a safra mais recente disponível (mês
corrente, -1, -2). Merge por mês/ano com a safra mais NOVA vencendo.

Flags: --upload publica no Blob; --no-backfill baixa só a safra mais recente
e faz merge com o blob publicado (modo cron).

Output: data-pipeline/out/fiscal-dbgg-fatores.json + Blob data/fiscal-dbgg-fatores.json
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import unicodedata
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

import requests
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from shared.blob_upload import maybe_upload_json  # noqa: E402
from shared.blob_download import download_json  # noqa: E402

DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/fiscal-dbgg-fatores.json"

XLSX_URL = (
    "https://www.bcb.gov.br/content/estatisticas/hist_estatisticasfiscais/"
    "{ym}_Tabelas_de_estatisticas_fiscais.xlsx"
)
UA = {"User-Agent": "Mozilla/5.0 (compatible; az-invest-dbgg-fatores/0.1)"}

SGS_URL = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados?formato=json"

SAFRA_MAIS_ANTIGA = 201805  # publicações anteriores retornam 404

TOL_SOMA_RS_MI = 1.0     # soma dos fatores vs variação da própria tabela
TOL_SGS_RS_MI = 5.0      # Δ SGS 13761 vs variação mensal
TOL_SGS_PP = 0.05        # Δ SGS 13762 (dez) vs variação anual em pp
TOL_SOMA_PP = 0.02       # identidade anual em pp (fatores + efeito PIB)

MESES = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Rótulos-chave (comparação sem acentos, minúscula). O rótulo da variação difere
# entre as tabelas (variação mensal vs var. acum. no ano) — o trecho comum basta.
LABELS = {
    "variacao": "divida bruta do gov. geral",   # + exigência extra de conter "var"
    "emissoes": "emissoes liquidas",
    "juros": "juros nominais",
    "cambio": "ajuste cambial",
    "externa_outros": "outros ajustes",
    "reconhecimento": "reconhecimento de dividas",
    "privatizacoes": "privatizacoes",
    "efeito_pib": "efeito crescimento",
}


def _norm(s) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _get(url, *, timeout=90, retries=4, sleep=4.0, ok404=False):
    last = None
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            if r.status_code == 404 and ok404:
                return None
            if r.status_code in (406, 429, 502, 503, 504):
                time.sleep((i + 1) * sleep)
                continue
            r.raise_for_status()
            return r
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep((i + 1) * 2)
    raise RuntimeError(f"falha apos {retries}: {url}: {last}")


def sgs_fetch_map(cod) -> dict[str, float]:
    """SGS mensal → {'YYYY-MM': valor}. Vazio em caso de falha (checks pulados com aviso)."""
    try:
        data = _get(SGS_URL.format(cod=cod)).json()
    except Exception as e:  # noqa: BLE001
        print(f"  [SGS {cod}] FALHA: {e}", file=sys.stderr)
        return {}
    out = {}
    for r in data:
        try:
            d, m, y = r["data"].split("/")
            out[f"{y}-{m}"] = float(str(r["valor"]).replace(",", "."))
        except Exception:  # noqa: BLE001
            continue
    print(f"  [SGS {cod}] {len(out)} pontos")
    return out


# ============================================================================
# Parser das Tabelas 18/19 — por rótulo, nunca índice fixo
# ============================================================================
def _find_rows(rows) -> dict[str, tuple]:
    """Localiza cada linha-chave pelo rótulo da coluna A. Aborta se faltar."""
    found = {}
    for row in rows:
        lab = _norm(row[0])
        if not lab:
            continue
        for key, frag in LABELS.items():
            if key in found:
                continue
            if frag in lab:
                if key == "variacao" and "var" not in lab.replace(frag, ""):
                    continue  # é a linha do SALDO, não da variação
                found[key] = row
    faltando = [k for k in LABELS if k not in found]
    if faltando:
        raise RuntimeError(f"rotulos nao encontrados na tabela: {faltando} — layout do BCB mudou, abortar")
    return found


def _columns(rows) -> list[tuple[int, int, int]]:
    """[(col_saldos, ano, mes)] — meses nos pares de colunas a partir da coluna C.

    Linha de anos = a que tem 'Discriminação' na coluna A; linha de meses = a
    primeira abaixo dela com nome de mês. Ano com forward-fill (a célula só é
    preenchida na primeira coluna de cada ano).
    """
    idx_disc = next(i for i, r in enumerate(rows) if _norm(r[0]).startswith("discriminacao"))
    row_anos = rows[idx_disc]
    row_meses = None
    for r in rows[idx_disc + 1: idx_disc + 6]:
        if any(_norm(c) in MESES for c in r[2:] if c):
            row_meses = r
            break
    if row_meses is None:
        raise RuntimeError("linha de meses nao encontrada — layout do BCB mudou, abortar")

    cols = []
    ano_atual = None
    for c in range(2, len(row_meses), 2):
        ano_cell = row_anos[c] if c < len(row_anos) else None
        if isinstance(ano_cell, (int, float)) and ano_cell:
            ano_atual = int(ano_cell)
        elif isinstance(ano_cell, str) and ano_cell.strip().isdigit():
            ano_atual = int(ano_cell.strip())
        mes = MESES.get(_norm(row_meses[c]))
        if mes is None or ano_atual is None:
            continue
        cols.append((c, ano_atual, mes))
    if not cols:
        raise RuntimeError("nenhuma coluna de mes reconhecida — layout do BCB mudou, abortar")
    return cols


def _val(row, c):
    v = row[c] if c < len(row) else None
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_tabela18(ws) -> dict[str, dict]:
    """Fluxos mensais em R$ milhões → {'YYYY-MM': {variacao, juros, emissoes, cambio, outros}}."""
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    linhas = _find_rows(rows)
    out = {}
    for c, ano, mes in _columns(rows):
        variacao = _val(linhas["variacao"], c)
        if variacao is None:
            continue
        emissoes = _val(linhas["emissoes"], c)
        juros = _val(linhas["juros"], c)
        cambio = _val(linhas["cambio"], c)
        outros = sum(_val(linhas[k], c) or 0.0 for k in ("externa_outros", "reconhecimento", "privatizacoes"))
        if None in (emissoes, juros, cambio):
            continue
        out[f"{ano}-{mes:02d}"] = {
            "variacao_dbgg_brl_mm": round(variacao, 2),
            "juros_nominais_brl_mm": round(juros, 2),
            "emissoes_liquidas_brl_mm": round(emissoes, 2),
            "ajuste_cambial_brl_mm": round(cambio, 2),
            "outros_brl_mm": round(outros, 2),
        }
    return out


def parse_tabela19(ws) -> dict[int, dict]:
    """Acumulado no ano — só colunas de DEZEMBRO (ano fechado), em pp do PIB.

    Também carrega os saldos R$ para a validação de identidade.
    """
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    linhas = _find_rows(rows)
    out = {}
    for c, ano, mes in _columns(rows):
        if mes != 12:
            continue
        variacao_pp = _val(linhas["variacao"], c + 1)
        if variacao_pp is None:
            continue
        emissoes_pp = _val(linhas["emissoes"], c + 1)
        juros_pp = _val(linhas["juros"], c + 1)
        cambio_pp = _val(linhas["cambio"], c + 1)
        efeito_pp = _val(linhas["efeito_pib"], c + 1)
        outros_pp = sum(_val(linhas[k], c + 1) or 0.0 for k in ("externa_outros", "reconhecimento", "privatizacoes"))
        if None in (emissoes_pp, juros_pp, cambio_pp, efeito_pp):
            continue
        rec = {
            "variacao_dbgg_pp_pib": round(variacao_pp, 4),
            "juros_nominais_pp": round(juros_pp, 4),
            "emissoes_liquidas_pp": round(emissoes_pp, 4),
            "ajuste_cambial_pp": round(cambio_pp, 4),
            "outros_pp": round(outros_pp, 4),
            "efeito_pib_pp": round(efeito_pp, 4),
        }
        # saldos R$ (para identidade) — não publicados no anual
        variacao_rs = _val(linhas["variacao"], c)
        soma_rs = None
        if variacao_rs is not None:
            soma_rs = sum(_val(linhas[k], c) or 0.0 for k in ("emissoes", "juros", "cambio", "externa_outros", "reconhecimento", "privatizacoes"))
        rec["_variacao_rs"] = variacao_rs
        rec["_soma_rs"] = soma_rs
        out[ano] = rec
    return out


def baixar_safra(ym: int, cache_dir: Path):
    """Baixa (com cache local) e parseia uma safra. None se a publicação não existe."""
    cache_dir.mkdir(parents=True, exist_ok=True)
    f = cache_dir / f"fiscais_{ym}.xlsx"
    if f.exists():
        content = f.read_bytes()
    else:
        r = _get(XLSX_URL.format(ym=ym), ok404=True)
        if r is None:
            return None
        content = r.content
        f.write_bytes(content)
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if "Tabela 18" not in wb.sheetnames or "Tabela 19" not in wb.sheetnames:
        raise RuntimeError(f"safra {ym}: Tabelas 18/19 ausentes do XLSX")
    mensal = parse_tabela18(wb["Tabela 18"])
    anual = parse_tabela19(wb["Tabela 19"])
    print(f"  [safra {ym}] mensal: {len(mensal)} meses ({min(mensal)}..{max(mensal)}); anual: {sorted(anual)}")
    return mensal, anual


def safras_backfill(hoje: date) -> list[int]:
    """Seed 201805 + {ano}01/{ano}06/{ano}11 de 2018 até hoje (ascendente)."""
    out = [SAFRA_MAIS_ANTIGA]
    ym_hoje = hoje.year * 100 + hoje.month
    for ano in range(2018, hoje.year + 1):
        for mes in (1, 6, 11):
            ym = ano * 100 + mes
            if SAFRA_MAIS_ANTIGA <= ym <= ym_hoje and ym not in out:
                out.append(ym)
    return sorted(out)


def safras_recentes(hoje: date) -> list[int]:
    """Mês corrente, -1, -2 (descendente) — a primeira que existir vence."""
    out = []
    y, m = hoje.year, hoje.month
    for _ in range(3):
        out.append(y * 100 + m)
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return out


# ============================================================================
# Validações (abortam o build/upload)
# ============================================================================
def validar(mensal: dict[str, dict], anual: dict[int, dict], *, sgs_13761, sgs_13762) -> list[str]:
    erros = []

    # 1) Identidade interna mensal: soma dos fatores = variação (tol. R$ 1 mi)
    for mes, r in mensal.items():
        soma = (r["juros_nominais_brl_mm"] + r["emissoes_liquidas_brl_mm"]
                + r["ajuste_cambial_brl_mm"] + r["outros_brl_mm"])
        if abs(soma - r["variacao_dbgg_brl_mm"]) > TOL_SOMA_RS_MI:
            erros.append(f"mensal {mes}: soma fatores R$ {soma:,.2f} mi != variacao R$ {r['variacao_dbgg_brl_mm']:,.2f} mi")

    # 2) Identidade interna anual em R$ e em pp
    for ano, r in anual.items():
        if r.get("_variacao_rs") is not None and r.get("_soma_rs") is not None:
            if abs(r["_soma_rs"] - r["_variacao_rs"]) > TOL_SOMA_RS_MI:
                erros.append(f"anual {ano}: soma fatores R$ {r['_soma_rs']:,.2f} mi != variacao R$ {r['_variacao_rs']:,.2f} mi")
        soma_pp = (r["juros_nominais_pp"] + r["emissoes_liquidas_pp"] + r["ajuste_cambial_pp"]
                   + r["outros_pp"] + r["efeito_pib_pp"])
        if abs(soma_pp - r["variacao_dbgg_pp_pib"]) > TOL_SOMA_PP:
            erros.append(f"anual {ano}: soma pp {soma_pp:.4f} != variacao {r['variacao_dbgg_pp_pib']:.4f} pp")

    # 3) Mensal vs SGS 13761 (DBGG saldos): Δ do estoque = variação da tabela
    if sgs_13761:
        n_check = 0
        for mes, r in sorted(mensal.items()):
            y, m = int(mes[:4]), int(mes[5:7])
            ant = f"{y - 1}-12" if m == 1 else f"{y}-{m - 1:02d}"
            if mes in sgs_13761 and ant in sgs_13761:
                delta = sgs_13761[mes] - sgs_13761[ant]
                n_check += 1
                if abs(delta - r["variacao_dbgg_brl_mm"]) > TOL_SGS_RS_MI:
                    erros.append(f"mensal {mes}: Δ SGS 13761 = R$ {delta:,.2f} mi != variacao R$ {r['variacao_dbgg_brl_mm']:,.2f} mi")
        print(f"  Check SGS 13761: {n_check} meses comparados")
    else:
        print("AVISO: SGS 13761 indisponivel — check mensal pulado", file=sys.stderr)

    # 4) Anual vs SGS 13762 (DBGG % PIB): Δ dez/ano - dez/ano-1 = variação em pp
    if sgs_13762:
        n_check = 0
        for ano, r in sorted(anual.items()):
            dez, dez_ant = f"{ano}-12", f"{ano - 1}-12"
            if dez in sgs_13762 and dez_ant in sgs_13762:
                delta = sgs_13762[dez] - sgs_13762[dez_ant]
                n_check += 1
                if abs(delta - r["variacao_dbgg_pp_pib"]) > TOL_SGS_PP:
                    erros.append(f"anual {ano}: Δ SGS 13762 = {delta:.4f} pp != variacao {r['variacao_dbgg_pp_pib']:.4f} pp")
        print(f"  Check SGS 13762: {n_check} anos comparados")
    else:
        print("AVISO: SGS 13762 indisponivel — check anual pulado", file=sys.stderr)

    return erros


# ============================================================================
# Main
# ============================================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--no-backfill", action="store_true",
                    help="so a safra mais recente + merge com o blob publicado (modo cron)")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = out_dir / "_cache_bcb_fiscais"

    hoje = date.today()
    mensal: dict[str, dict] = {}
    anual: dict[int, dict] = {}

    if args.no_backfill:
        # Base = blob publicado; por cima, a safra mais recente disponível.
        blob = download_json(BLOB_PATH)
        if blob:
            for r in blob.get("mensal") or []:
                mensal[r["data"]] = {k: v for k, v in r.items() if k != "data"}
            for r in blob.get("anual") or []:
                anual[int(r["ano"])] = {k: v for k, v in r.items() if k != "ano"}
            print(f"  Blob existente: {len(mensal)} meses / {len(anual)} anos")
        else:
            print("AVISO: blob inexistente — no-backfill vai publicar so a safra recente", file=sys.stderr)
        achou = False
        for ym in safras_recentes(hoje):
            res = baixar_safra(ym, cache_dir)
            if res is None:
                print(f"  [safra {ym}] nao publicada ainda")
                continue
            m, a = res
            mensal.update(m)
            for ano, rec in a.items():
                anual[ano] = rec
            achou = True
            break
        if not achou:
            print("ERRO: nenhuma safra recente disponivel (mes corrente, -1, -2)", file=sys.stderr)
            sys.exit(1)
    else:
        # Backfill completo: safra mais NOVA vence (iteracao ascendente + recentes no fim)
        lista = safras_backfill(hoje)
        recentes_asc = [ym for ym in sorted(safras_recentes(hoje)) if ym not in lista]
        for ym in lista + recentes_asc:
            res = baixar_safra(ym, cache_dir)
            if res is None:
                print(f"  [safra {ym}] 404 — pulada")
                continue
            m, a = res
            mensal.update(m)
            for ano, rec in a.items():
                anual[ano] = rec

    if not mensal or not anual:
        print("ERRO: nada parseado", file=sys.stderr)
        sys.exit(1)

    # === Validações (abortam) ===
    print("Validando identidades e SGS...")
    sgs_13761 = sgs_fetch_map(13761)
    sgs_13762 = sgs_fetch_map(13762)
    erros = validar(mensal, anual, sgs_13761=sgs_13761, sgs_13762=sgs_13762)
    if erros:
        print(f"ERRO: {len(erros)} violacoes de validacao — upload abortado:", file=sys.stderr)
        for e in erros[:30]:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(1)

    payload = {
        "schema_version": 1,
        "gerado_em": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "_fonte": (
            "BCB — Tabelas de estatísticas fiscais (XLSX das Notas Econômico-Financeiras, "
            "hist_estatisticasfiscais): Tabela 18 (fatores condicionantes da DBGG, fluxos mensais, "
            "R$ milhões) e Tabela 19 (acumulado no ano, % PIB). Reconstrução histórica iterando "
            "as safras de publicação (mais antiga disponível: 201805); a safra mais nova vence no merge. "
            "Validado contra SGS 13761 (DBGG saldos) e 13762 (DBGG % PIB)."
        ),
        "_nota": (
            "Decomposição da variação da DBGG: emissões líquidas + juros nominais + ajuste cambial + "
            "outros (dívida externa-outros ajustes + reconhecimento de dívidas + privatizações). "
            "No anual, em pontos do PIB, entra também o efeito do crescimento do PIB nominal sobre o "
            "denominador: juros + emissões + câmbio + outros + efeito PIB ≈ variação da razão DBGG/PIB. "
            "Mensal em R$ milhões (sem efeito PIB, que só existe na razão)."
        ),
        "anual": [
            {"ano": ano, **{k: v for k, v in rec.items() if not k.startswith("_")}}
            for ano, rec in sorted(anual.items())
        ],
        "mensal": [
            {"data": mes, **rec} for mes, rec in sorted(mensal.items())
        ],
    }

    out_file = out_dir / "fiscal-dbgg-fatores.json"
    out_file.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"  -> {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")
    print(f"  Mensal: {len(payload['mensal'])} meses ({payload['mensal'][0]['data']} -> {payload['mensal'][-1]['data']})")
    print(f"  Anual: {len(payload['anual'])} anos ({payload['anual'][0]['ano']} -> {payload['anual'][-1]['ano']})")
    ult = payload["mensal"][-1]
    print(f"  Último mês {ult['data']}: variação R$ {ult['variacao_dbgg_brl_mm']:,.0f} mi (juros {ult['juros_nominais_brl_mm']:,.0f})")

    if args.upload:
        maybe_upload_json(out_file, BLOB_PATH)


if __name__ == "__main__":
    main()
