"""Build do JSON do Painel Fiscal — clássicos brasileiros (receita, gastos, dívida).

Fontes:
- BCB SGS (DBGG, DLSP, primário/juros/NFSP % PIB, REER, Selic, IPCA, PIB nominal, PIB real)
- BCB Olinda (Focus — Selic, IPCA, PIB, Câmbio)
- Tesouro Nacional RTN (XLSX): receita líquida, despesa primária, juros nominais do
  GOVERNO CENTRAL — séries mensais R$ MM desde 1997

Output: data-pipeline/out/fiscal-classicos.json + upload Blob em data/fiscal-classicos.json

Convenção contábil:
- Primário positivo = SUPERÁVIT (convenção STN/BCB)
- NFSP positivo = DÉFICIT (oposto do primário) — convenção BCB SGS
- Juros nominais no RTN vêm negativos (saídas); convertemos pra positivo (custo)
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from shared.blob_upload import maybe_upload_json  # noqa: E402

DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/fiscal-classicos.json"

UA = {"User-Agent": "Mozilla/5.0 (compatible; az-invest-fiscal/0.2)"}


def _get(url, *, timeout=60, retries=4, sleep=4.0):
    last = None
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            if r.status_code in (406, 429, 502, 503, 504):
                time.sleep((i + 1) * sleep)
                continue
            r.raise_for_status()
            return r
        except Exception as e:
            last = e
            time.sleep((i + 1) * 2)
    raise RuntimeError(f"falha apos {retries}: {last}")


def _to_float(v):
    if v in ("", "-", "..", "...", None):
        return None
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _parse_sgs(s, daily=False):
    d, m, y = s.split("/")
    return f"{y}-{m}-{d}" if daily else f"{y}-{m}"


SGS_URL = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados?formato=json"
SGS_URL_FROM = SGS_URL + "&dataInicial={inicio}"


def sgs_fetch(cod, *, daily=False, since=None):
    url = SGS_URL_FROM.format(cod=cod, inicio=since) if since else SGS_URL.format(cod=cod)
    print(f"  [SGS {cod}]")
    try:
        data = _get(url).json()
    except Exception as e:
        print(f"  [SGS {cod}] FALHA: {e}", file=sys.stderr)
        return []
    out = []
    for r in data:
        try:
            out.append({"data": _parse_sgs(r["data"], daily), "valor": _to_float(r["valor"])})
        except Exception:
            continue
    return out


# Tesouro RTN — XLSX dinâmico do SISWEB
RTN_URL = "http://sisweb.tesouro.gov.br/apex/cosis/thot/link/rtn/serie-historica?conteudo=cdn"

RTN_LINHAS = {
    "receita_total": 6,
    "transferencias": 29,
    "receita_liquida": 38,
    "despesa_total": 39,
    "previdencia": 40,
    "pessoal": 41,
    "outras_obrigatorias": 42,
    "abono_seguro": 43,           # 4.3.01 Abono e Seguro Desemprego
    "bpc_loas": 47,               # 4.3.05 Benefícios LOAS/RMV
    "fundeb": 52,                 # 4.3.10 FUNDEB
    "subsidios": 57,              # 4.3.15 Subsídios, subvenções, Proagro
    "discricionarias": 65,
    "primario_acima": 66,
    "juros_nominais": 74,
    "nominal": 75,
    # Receita administrada RFB por tributo (linhas 8-17 do RTN aba 1.1)
    "imposto_importacao": 8,         # 1.1.01 Imposto sobre Importação
    "ipi": 9,                        # 1.1.02 IPI
    "imposto_renda": 10,             # 1.1.03 Imposto de Renda
    "iof": 11,                       # 1.1.04 IOF
    "cofins": 12,                    # 1.1.05 Cofins
    "pis_pasep": 13,                 # 1.1.06 PIS/Pasep
    "csll": 14,                      # 1.1.07 CSLL
    "cide_combustiveis": 16,         # 1.1.09 CIDE-Combustíveis
    # Arrecadacao previdenciaria (RGPS)
    "rgps_arrecadacao": 19,          # 1.3 Arrecadação Líquida para RGPS
    # Receita nao-administrada
    "concessoes": 21,                # 1.4.1 Concessões e Permissões
    "dividendos": 22,                # 1.4.2 Dividendos e Participações
    "recursos_naturais": 24,         # 1.4.4 Receitas de Exploração de Recursos Naturais
    # ── v2: agregados de FAMÍLIA da receita (stack completo que fecha com a receita total)
    "receita_administrada_rfb": 7,   # 1.1 Receita Administrada pela RFB (agregado)
    "incentivos_fiscais": 18,        # 1.2 Incentivos Fiscais (negativo, pequeno)
    "receita_nao_administrada": 20,  # 1.4 Receitas Não Administradas pela RFB (agregado)
    # ── v2: abertura da despesa 4.4 (rigidez orçamentária)
    "despesa_prog_financeira": 63,   # 4.4 Despesas do Executivo Sujeitas à Programação Financeira (agregado)
    "obrig_controle_fluxo": 64,      # 4.4.1 Obrigatórias com Controle de Fluxo
}

# v2: validação textual — o RTN usa linhas fixas; se o Tesouro inserir uma linha o
# parser leria IPI como IR SEM NENHUM ERRO. O rótulo da coluna A precisa conter o
# trecho esperado (comparação sem acentos), senão o build aborta antes de publicar.
RTN_LABELS = {
    "receita_total": "1. RECEITA TOTAL",
    "transferencias": "2. TRANSF",
    "receita_liquida": "3. RECEITA LIQUIDA",
    "despesa_total": "4. DESPESA TOTAL",
    "previdencia": "4.1",
    "pessoal": "4.2",
    "outras_obrigatorias": "4.3",
    "abono_seguro": "4.3.01",
    "bpc_loas": "4.3.05",
    "fundeb": "4.3.10",
    "subsidios": "4.3.15",
    "discricionarias": "4.4.2",
    "primario_acima": "5. RESULTADO PRIMARIO",
    "juros_nominais": "9. JUROS NOMINAIS",
    "nominal": "10. RESULTADO NOMINAL",
    "imposto_importacao": "1.1.01",
    "ipi": "1.1.02",
    "imposto_renda": "1.1.03",
    "iof": "1.1.04",
    "cofins": "1.1.05",
    "pis_pasep": "1.1.06",
    "csll": "1.1.07",
    "cide_combustiveis": "1.1.09",
    "rgps_arrecadacao": "1.3",
    "concessoes": "1.4.1",
    "dividendos": "1.4.2",
    "recursos_naturais": "1.4.4",
    "receita_administrada_rfb": "1.1 -",
    "incentivos_fiscais": "1.2",
    "receita_nao_administrada": "1.4 -",
    "despesa_prog_financeira": "4.4",
    "obrig_controle_fluxo": "4.4.1",
}


def _normaliza_rotulo(s):
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).upper().strip()


def baixa_rtn_xlsx():
    print(f"  [Tesouro RTN] baixando")
    r = _get(RTN_URL, timeout=60)
    return BytesIO(r.content)


def parse_rtn(xlsx_stream):
    wb = load_workbook(xlsx_stream, data_only=True, read_only=True)
    sh = wb["1.1"]
    header = next(sh.iter_rows(min_row=5, max_row=5, values_only=True))
    datas_idx = []
    for i, h in enumerate(header[1:], 1):
        if h:
            try:
                if hasattr(h, "year"):
                    datas_idx.append((i, f"{h.year:04d}-{h.month:02d}"))
                else:
                    datas_idx.append((i, str(h)[:7]))
            except Exception:
                continue

    series = {k: [] for k in RTN_LINHAS}
    rotulos_errados = []
    for chave, row_num in RTN_LINHAS.items():
        row = next(sh.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        # v2: valida o rótulo da coluna A — se o Tesouro inserir/mover uma linha,
        # abortamos em vez de publicar IPI rotulado como IR.
        esperado = RTN_LABELS.get(chave)
        if esperado and esperado.upper() not in _normaliza_rotulo(row[0]):
            rotulos_errados.append(f"linha {row_num} ({chave}): esperado '{esperado}', achado '{str(row[0])[:60]}'")
            continue
        for i, mes in datas_idx:
            v = row[i] if i < len(row) else None
            if v in ("", None):
                continue
            try:
                vf = float(v)
                if chave == "juros_nominais":
                    vf = -vf  # flip pra positivo (custo)
                series[chave].append({"data": mes, "valor": round(vf, 2)})
            except (TypeError, ValueError):
                continue
    if rotulos_errados:
        print("[ERROR] RTN: rotulos fora do lugar — layout do XLSX mudou; abortando:", file=sys.stderr)
        for r in rotulos_errados:
            print(f"  {r}", file=sys.stderr)
        sys.exit(2)
    return series


# Focus
FOCUS_BASE = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata"


def focus_anuais(indicador, ano_atual):
    indicador_url = indicador.replace(" ", "%20")
    url = (
        f"{FOCUS_BASE}/ExpectativasMercadoAnuais?$format=json&$top=20000"
        f"&$filter=Indicador%20eq%20%27{indicador_url}%27%20and%20Data%20ge%20%27{ano_atual - 1}-01-01%27"
        f"&$orderby=Data%20desc"
    )
    print(f"  [Focus {indicador}]")
    try:
        data = _get(url, timeout=90).json().get("value", [])
    except Exception as e:
        print(f"  [Focus {indicador}] FALHA: {e}", file=sys.stderr)
        return {}
    out = {}
    for r in data:
        try:
            ano = int(r["DataReferencia"])
        except (KeyError, ValueError):
            continue
        if ano not in (ano_atual, ano_atual + 1, ano_atual + 2, ano_atual + 3):
            continue
        out.setdefault(ano, []).append({
            "data": r.get("Data", "")[:10],
            "mediana": _to_float(r.get("Mediana")),
            "media": _to_float(r.get("Media")),
            "dp": _to_float(r.get("DesvioPadrao")),
            "min": _to_float(r.get("Minimo")),
            "max": _to_float(r.get("Maximo")),
        })
    for ano in out:
        out[ano].sort(key=lambda x: x["data"])
        out[ano] = out[ano][-365:]
    return out


def soma_12m(serie):
    if len(serie) < 12:
        return []
    serie = sorted(serie, key=lambda x: x["data"])
    out = []
    for i in range(11, len(serie)):
        window = serie[i - 11:i + 1]
        vals = [w["valor"] for w in window if w["valor"] is not None]
        if len(vals) < 12:
            continue
        out.append({"data": serie[i]["data"], "valor_12m": round(sum(vals), 2)})
    return out


def divide_por_pib(serie_12m, pib_map):
    if not pib_map:
        return []
    meses_pib = sorted(pib_map.keys())
    out = []
    ultimo_pib = None
    cur_idx = 0
    for r in sorted(serie_12m, key=lambda x: x["data"]):
        while cur_idx < len(meses_pib) and meses_pib[cur_idx] <= r["data"]:
            ultimo_pib = pib_map[meses_pib[cur_idx]]
            cur_idx += 1
        if ultimo_pib is None or ultimo_pib == 0:
            continue
        out.append({"data": r["data"], "valor_pct": round(r["valor_12m"] / ultimo_pib * 100, 4)})
    return out


def divide_por_receita(serie_12m, receita_12m):
    rmap = {r["data"]: r["valor_12m"] for r in receita_12m}
    out = []
    for r in serie_12m:
        rec = rmap.get(r["data"])
        if rec is None or rec == 0:
            continue
        out.append({"data": r["data"], "valor_pct": round(r["valor_12m"] / rec * 100, 4)})
    return out


def selic_real_ex_post(selic_diaria, ipca_mensal):
    selic_por_mes = {}
    for r in selic_diaria:
        if r["valor"] is None:
            continue
        selic_por_mes[r["data"][:7]] = r["valor"]
    out = []
    for r in ipca_mensal:
        if r["valor"] is None:
            continue
        s = selic_por_mes.get(r["data"])
        if s is None:
            continue
        real = ((1 + s / 100) / (1 + r["valor"] / 100) - 1) * 100
        out.append({"data": r["data"], "selic_nominal_pct": s, "ipca_12m_pct": r["valor"], "selic_real_pct": round(real, 4)})
    return out


def pib_real_yoy(pib_real_idx):
    out = []
    idx_map = {r["data"]: r["valor"] for r in pib_real_idx if r["valor"] is not None}
    meses = sorted(idx_map.keys())
    for mes in meses:
        y, m = mes.split("-")
        ant = f"{int(y) - 1}-{m}"
        if ant in idx_map and idx_map[ant]:
            yoy = (idx_map[mes] / idx_map[ant] - 1) * 100
            out.append({"data": mes, "valor_yoy_pct": round(yoy, 4)})
    return out


def last_val(serie, key="valor"):
    for r in reversed(serie):
        v = r.get(key)
        if v is not None:
            return r
    return None


# ── helpers v2 ───────────────────────────────────────────────────────────────

def compoe_indice_ipca(ipca_mensal_var):
    """SGS 433 (variação % mensal) → número-índice composto, NORMALIZADO para o
    último mês = 100 ("a preços de hoje"). O 433 NÃO é número-índice — compor é o
    jeito certo de deflacionar mês a mês. A normalização importa: composto desde
    1980 o índice atravessa a hiperinflação e os valores deflacionados ficariam
    microscópicos (o round das somas 12m os zeraria)."""
    out = {}
    idx = 100.0
    for r in sorted(ipca_mensal_var, key=lambda x: x["data"]):
        if r["valor"] is None:
            continue
        idx *= 1 + r["valor"] / 100
        out[r["data"]] = idx
    if not out:
        return out
    base = out[sorted(out.keys())[-1]]
    return {d: v / base * 100 for d, v in out.items()}


def real_12m_yoy(serie_mensal_nominal, indice_ipca):
    """Crescimento REAL YoY do acumulado 12m, deflacionando MÊS A MÊS pelo índice
    (dividir a razão nominal pelo IPCA YoY do endpoint distorce quando a inflação
    muda dentro da janela)."""
    reais = []
    for r in sorted(serie_mensal_nominal, key=lambda x: x["data"]):
        idx = indice_ipca.get(r["data"])
        if r["valor"] is None or not idx:
            continue
        reais.append({"data": r["data"], "valor": r["valor"] / idx * 100})
    soma = soma_12m(reais)
    smap = {r["data"]: r["valor_12m"] for r in soma}
    out = []
    for data, v in sorted(smap.items()):
        y, m = data.split("-")
        ant = f"{int(y) - 1}-{m}"
        if smap.get(ant):
            out.append({"data": data, "valor_yoy_pct": round((v / smap[ant] - 1) * 100, 2)})
    return out


def serie_mensal_e_ytd(serie_mensal):
    """Acumulado jan→mês por ano-calendário (acompanhamento da meta como o RTN publica)."""
    ytd = {}
    acum = {}
    for r in sorted(serie_mensal, key=lambda x: x["data"]):
        if r["valor"] is None:
            continue
        ano, mes = r["data"].split("-")
        acum[ano] = acum.get(ano, 0.0) + r["valor"]
        ytd.setdefault(ano, []).append({"mes": int(mes), "acum_brl_mm": round(acum[ano], 2)})
    return ytd


def calcula_sustentabilidade(juros_sp_pct, dlsp_pct, primario_sp_pct, pib_map):
    """r − g e primário estabilizador com PERÍMETRO ÚNICO (setor público consolidado):
    r = taxa implícita da DLSP = juros nominais 12m (R$) ÷ DLSP média dos 12m (R$);
    g = PIB nominal acumulado 12m YoY (SGS 4382);
    p* = (r−g)/(1+g) × DLSP%PIB de 12 meses atrás (Blanchard, horizonte anual).
    Juros consolidados sobre DBGG (governo geral) seria perímetro misto — o vício
    que este v2 elimina. Calculado UMA vez aqui; o front nunca recalcula."""
    juros_map = {r["data"]: r["valor"] for r in juros_sp_pct if r["valor"] is not None}
    dlsp_map = {r["data"]: r["valor"] for r in dlsp_pct if r["valor"] is not None}
    prim_map = {r["data"]: r["valor_pct"] for r in primario_sp_pct if r["valor_pct"] is not None}
    meses = sorted(m for m in juros_map if m in dlsp_map and m in pib_map)
    out = []
    for t in meses:
        y, m = t.split("-")
        t12 = f"{int(y) - 1}-{m}"
        pib_t, pib_t12 = pib_map.get(t), pib_map.get(t12)
        if not pib_t or not pib_t12:
            continue
        g = pib_t / pib_t12 - 1
        # estoque médio da DLSP em R$ nos 12 meses encerrados em t
        janela = []
        ano_i, mes_i = int(y), int(m)
        for k in range(12):
            mm = mes_i - k
            yy = ano_i
            while mm <= 0:
                mm += 12
                yy -= 1
            chave = f"{yy:04d}-{mm:02d}"
            if dlsp_map.get(chave) and pib_map.get(chave):
                janela.append(dlsp_map[chave] / 100 * pib_map[chave])
        if len(janela) < 12:
            continue
        dlsp_medio_rs = sum(janela) / len(janela)
        juros_rs = juros_map[t] / 100 * pib_t
        r_aa = juros_rs / dlsp_medio_rs
        dlsp_t12 = dlsp_map.get(t12)
        p_estab = ((r_aa - g) / (1 + g) * dlsp_t12) if dlsp_t12 else None
        out.append({
            "data": t,
            "r_aa_pct": round(r_aa * 100, 2),
            "g_aa_pct": round(g * 100, 2),
            "r_menos_g_pp": round((r_aa - g) * 100, 2),
            "primario_estabilizador_pct_pib": round(p_estab, 2) if p_estab is not None else None,
            "primario_realizado_sp_pct_pib": prim_map.get(t),
            "dlsp_pct_pib": dlsp_map.get(t),
        })
    return out


def decompoe_dlsp_anual(sust_serie, dlsp_pct, juros_sp_pct, primario_sp_pct):
    """Por que a dívida (líquida, consolidado) subiu? Δb anual ≈ juros − primário −
    efeito crescimento (g/(1+g)·b_ini) + resíduo (ajustes patrimoniais/cambiais).
    Identidade limpa só no consolidado — a decomposição OFICIAL da DBGG (fatores
    condicionantes da Nota de Imprensa) entra quando a coleta for adicionada."""
    dlsp_map = {r["data"]: r["valor"] for r in dlsp_pct if r["valor"] is not None}
    juros_map = {r["data"]: r["valor"] for r in juros_sp_pct if r["valor"] is not None}
    prim_map = {r["data"]: r["valor_pct"] for r in primario_sp_pct if r["valor_pct"] is not None}
    g_map = {r["data"]: r["g_aa_pct"] for r in sust_serie}
    anos = sorted({d[:4] for d in dlsp_map})
    out = []
    for ano in anos:
        dez, dez_ant = f"{ano}-12", f"{int(ano) - 1}-12"
        b1, b0 = dlsp_map.get(dez), dlsp_map.get(dez_ant)
        j, p, g = juros_map.get(dez), prim_map.get(dez), g_map.get(dez)
        if None in (b1, b0, j, p, g):
            continue
        efeito_pib = -(g / 100) / (1 + g / 100) * b0
        residuo = (b1 - b0) - (j - p + efeito_pib)
        out.append({
            "ano": ano,
            "delta_pp": round(b1 - b0, 2),
            "juros_pp": round(j, 2),
            "primario_pp": round(-p, 2),  # superávit REDUZ a dívida (sinal já aplicado p/ empilhar)
            "efeito_crescimento_pp": round(efeito_pib, 2),
            "residuo_pp": round(residuo, 2),
            "dlsp_fim_pct_pib": b1,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--no-merge", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "fiscal-classicos.json"

    print("== [1/4] Tesouro RTN (XLSX) ==")
    rtn_data = parse_rtn(baixa_rtn_xlsx())
    print(f"  Receita liquida: {len(rtn_data['receita_liquida'])} obs")

    print("\n== [2/4] BCB SGS diarias ==")
    reservas_diaria = sgs_fetch(13621, daily=True, since="01/01/2018")
    time.sleep(2)
    selic_diaria = sgs_fetch(1178, daily=True, since="01/01/2018")
    time.sleep(2)

    print("\n== [3/4] BCB SGS mensais ==")
    series_mensal = {
        "dbgg": 13762, "dlsp_total": 4513, "dlsp_gov_central": 4503,
        "nfsp_sp": 5727, "nfsp_central": 5717,
        "juros_sp_pct": 5718, "juros_central_pct": 5728,
        "pib_12m_brl": 4382, "reer": 11752,
        "ipca_12m": 13522, "pib_real_idx": 22099,
        # Composicao DPMFi por indexador (BCB SGS)
        "dpmfi_pct_selic": 4177,        # % indexado a Selic/LFT
        "dpmfi_pct_prefixado": 4178,    # % prefixado (LTN+NTN-F)
        "dpmfi_pct_cambio": 4175,       # % cambial
        "dpmfi_pct_tr": 4174,           # % TR
        "dpmfi_pct_outros": 4176,       # % outros indexadores
        "dpmfi_pct_indices_precos": 12001,  # % índices de preços (NTN-B etc.) — fecha o stack em ~100
        "ipca_mensal_var": 433,         # IPCA variação % MENSAL (p/ compor número-índice e deflacionar)
        # Credito total economia (saldo % PIB)
        "credito_total_pct_pib": 20622, # Saldo total de credito / PIB
        # Selic over diaria? Ja temos 1178
    }
    sgs = {}
    for nome, cod in series_mensal.items():
        sgs[nome] = sgs_fetch(cod)
        time.sleep(0.4)

    print("\n== [4/4] Derivados ==")
    reservas_mensal = {}
    for r in reservas_diaria:
        if r["valor"] is None:
            continue
        reservas_mensal[r["data"][:7]] = r["valor"]
    reservas_mensal = [{"data": k, "valor": v} for k, v in sorted(reservas_mensal.items())]

    pib_map = {r["data"]: r["valor"] for r in sgs["pib_12m_brl"] if r["valor"] is not None}

    receita_liquida_12m = soma_12m(rtn_data["receita_liquida"])
    despesa_total_12m = soma_12m(rtn_data["despesa_total"])
    primario_central_12m = soma_12m(rtn_data["primario_acima"])
    juros_central_12m = soma_12m(rtn_data["juros_nominais"])
    previdencia_12m = soma_12m(rtn_data["previdencia"])
    pessoal_12m = soma_12m(rtn_data["pessoal"])
    outras_obrig_12m = soma_12m(rtn_data["outras_obrigatorias"])
    discricionarias_12m = soma_12m(rtn_data["discricionarias"])
    # Receita por tributo
    ii_12m = soma_12m(rtn_data["imposto_importacao"])
    ipi_12m = soma_12m(rtn_data["ipi"])
    ir_12m = soma_12m(rtn_data["imposto_renda"])
    iof_12m = soma_12m(rtn_data["iof"])
    cofins_12m = soma_12m(rtn_data["cofins"])
    pis_12m = soma_12m(rtn_data["pis_pasep"])
    csll_12m = soma_12m(rtn_data["csll"])
    cide_12m = soma_12m(rtn_data["cide_combustiveis"])
    rgps_12m = soma_12m(rtn_data["rgps_arrecadacao"])
    concessoes_12m = soma_12m(rtn_data["concessoes"])
    dividendos_12m = soma_12m(rtn_data["dividendos"])
    rec_naturais_12m = soma_12m(rtn_data["recursos_naturais"])
    abono_seguro_12m = soma_12m(rtn_data["abono_seguro"])
    bpc_loas_12m = soma_12m(rtn_data["bpc_loas"])
    fundeb_12m = soma_12m(rtn_data["fundeb"])
    subsidios_12m = soma_12m(rtn_data["subsidios"])

    receita_pct_pib = divide_por_pib(receita_liquida_12m, pib_map)
    despesa_pct_pib = divide_por_pib(despesa_total_12m, pib_map)
    primario_central_pct_pib = divide_por_pib(primario_central_12m, pib_map)
    juros_central_pct_pib = divide_por_pib(juros_central_12m, pib_map)
    previdencia_pct_pib = divide_por_pib(previdencia_12m, pib_map)
    pessoal_pct_pib = divide_por_pib(pessoal_12m, pib_map)

    despesa_pct_rec = divide_por_receita(despesa_total_12m, receita_liquida_12m)
    juros_pct_rec = divide_por_receita(juros_central_12m, receita_liquida_12m)
    primario_pct_rec = divide_por_receita(primario_central_12m, receita_liquida_12m)
    previdencia_pct_rec = divide_por_receita(previdencia_12m, receita_liquida_12m)
    pessoal_pct_rec = divide_por_receita(pessoal_12m, receita_liquida_12m)

    # Decomposicao receita por tributo (% PIB)
    ii_pct_pib = divide_por_pib(ii_12m, pib_map)
    ipi_pct_pib = divide_por_pib(ipi_12m, pib_map)
    ir_pct_pib = divide_por_pib(ir_12m, pib_map)
    iof_pct_pib = divide_por_pib(iof_12m, pib_map)
    cofins_pct_pib = divide_por_pib(cofins_12m, pib_map)
    pis_pct_pib = divide_por_pib(pis_12m, pib_map)
    csll_pct_pib = divide_por_pib(csll_12m, pib_map)
    cide_pct_pib = divide_por_pib(cide_12m, pib_map)
    rgps_pct_pib = divide_por_pib(rgps_12m, pib_map)
    concessoes_pct_pib = divide_por_pib(concessoes_12m, pib_map)
    dividendos_pct_pib = divide_por_pib(dividendos_12m, pib_map)
    rec_naturais_pct_pib = divide_por_pib(rec_naturais_12m, pib_map)

    # Decomposicao expandida de despesa (% PIB)
    abono_seguro_pct_pib = divide_por_pib(abono_seguro_12m, pib_map)
    bpc_pct_pib = divide_por_pib(bpc_loas_12m, pib_map)
    fundeb_pct_pib = divide_por_pib(fundeb_12m, pib_map)
    subsidios_pct_pib = divide_por_pib(subsidios_12m, pib_map)
    discricionarias_pct_pib = divide_por_pib(discricionarias_12m, pib_map)
    outras_obrig_pct_pib = divide_por_pib(outras_obrig_12m, pib_map)

    # Decomposicao expandida (% Receita liquida)
    abono_seguro_pct_rec = divide_por_receita(abono_seguro_12m, receita_liquida_12m)
    bpc_pct_rec = divide_por_receita(bpc_loas_12m, receita_liquida_12m)
    fundeb_pct_rec = divide_por_receita(fundeb_12m, receita_liquida_12m)
    subsidios_pct_rec = divide_por_receita(subsidios_12m, receita_liquida_12m)
    discricionarias_pct_rec = divide_por_receita(discricionarias_12m, receita_liquida_12m)
    outras_obrig_pct_rec = divide_por_receita(outras_obrig_12m, receita_liquida_12m)

    selic_real = selic_real_ex_post(selic_diaria, sgs["ipca_12m"])
    pib_real_yoy_serie = pib_real_yoy(sgs["pib_real_idx"])

    # ── v2: famílias de receita (stack que FECHA com a receita total) ──
    admin_12m = soma_12m(rtn_data["receita_administrada_rfb"])
    incentivos_12m = soma_12m(rtn_data["incentivos_fiscais"])
    nao_admin_12m = soma_12m(rtn_data["receita_nao_administrada"])
    admin_pct_pib = divide_por_pib(admin_12m, pib_map)
    incentivos_pct_pib = divide_por_pib(incentivos_12m, pib_map)
    nao_admin_pct_pib = divide_por_pib(nao_admin_12m, pib_map)
    receita_total_12m = soma_12m(rtn_data["receita_total"])
    # validação de identidade: 1.1 + 1.2 + 1.3 + 1.4 ≈ receita total (12m, último mês)
    if admin_12m and receita_total_12m:
        u = receita_total_12m[-1]["data"]
        partes = 0.0
        for s in (admin_12m, incentivos_12m, soma_12m(rtn_data["rgps_arrecadacao"]), nao_admin_12m):
            m = {r["data"]: r["valor_12m"] for r in s}
            if m.get(u) is None:
                partes = None
                break
            partes += m[u]
        total = receita_total_12m[-1]["valor_12m"]
        if partes is not None and total and abs(partes - total) / total > 0.005:
            print(f"[WARN] famílias da receita não fecham com o total em {u}: {partes:.0f} vs {total:.0f}", file=sys.stderr)

    # v2: dividendos + concessões NUMA série (o front rotulava 'Dividendos+Concessões' plotando só dividendos)
    div_map = {r["data"]: r["valor"] for r in rtn_data["dividendos"] if r["valor"] is not None}
    div_conc_mensal = []
    for r in rtn_data["concessoes"]:
        if r["valor"] is None or div_map.get(r["data"]) is None:
            continue
        div_conc_mensal.append({"data": r["data"], "valor": round(r["valor"] + div_map[r["data"]], 2)})
    dividendos_concessoes_pct_pib = divide_por_pib(soma_12m(div_conc_mensal), pib_map)

    # ── v2: rubrica residual 'demais obrigatórias' (4.3 contém abono/BPC/FUNDEB/subsídios;
    # somar o 4.3 inteiro ao stack DUPLICARIA ~2pp — o residual é a 8ª fatia correta) ──
    subs_maps = [
        {r["data"]: r["valor"] for r in rtn_data[k] if r["valor"] is not None}
        for k in ("abono_seguro", "bpc_loas", "fundeb", "subsidios")
    ]
    demais_obrig_mensal = []
    for r in rtn_data["outras_obrigatorias"]:
        if r["valor"] is None:
            continue
        subs = [m.get(r["data"]) for m in subs_maps]
        if any(v is None for v in subs):
            continue
        demais_obrig_mensal.append({"data": r["data"], "valor": round(r["valor"] - sum(subs), 2)})
    demais_obrig_pct_pib = divide_por_pib(soma_12m(demais_obrig_mensal), pib_map)
    obrig_fluxo_pct_pib = divide_por_pib(soma_12m(rtn_data["obrig_controle_fluxo"]), pib_map)

    # validação de identidade da despesa: 4.1+4.2+4.3+4.4 ≈ despesa total (último 12m)
    desp_partes_12m = [soma_12m(rtn_data[k]) for k in ("previdencia", "pessoal", "outras_obrigatorias", "despesa_prog_financeira")]
    if despesa_total_12m and all(desp_partes_12m):
        u = despesa_total_12m[-1]["data"]
        soma_partes = 0.0
        for s in desp_partes_12m:
            m = {r["data"]: r["valor_12m"] for r in s}
            if m.get(u) is None:
                soma_partes = None
                break
            soma_partes += m[u]
        total = despesa_total_12m[-1]["valor_12m"]
        if soma_partes is not None and total and abs(soma_partes - total) / total > 0.005:
            print(f"[WARN] rubricas da despesa não fecham com o total em {u}: {soma_partes:.0f} vs {total:.0f}", file=sys.stderr)

    # ── v2: arcabouço — crescimento REAL deflacionado mês a mês pelo índice composto do 433 ──
    indice_ipca = compoe_indice_ipca(sgs["ipca_mensal_var"])
    despesa_real_yoy = real_12m_yoy(rtn_data["despesa_total"], indice_ipca)
    receita_real_yoy = real_12m_yoy(rtn_data["receita_liquida"], indice_ipca)

    # ── v2: acompanhamento da meta pelo ACUMULADO NO ANO (como o RTN publica) ──
    primario_ytd = serie_mensal_e_ytd(rtn_data["primario_acima"])

    ano_atual = datetime.now(timezone.utc).year
    focus_selic = focus_anuais("Selic", ano_atual)
    focus_ipca = focus_anuais("IPCA", ano_atual)
    focus_pib = focus_anuais("PIB Total", ano_atual)
    focus_cambio = focus_anuais("Câmbio", ano_atual)

    pib_12m_recente = sgs["pib_12m_brl"][-1]["valor"] if sgs["pib_12m_brl"] else None

    nominal_sp_pct = [{"data": r["data"], "valor_pct": -r["valor"] if r["valor"] is not None else None} for r in sgs["nfsp_sp"]]
    primario_sp_pct = []
    juros_sp_map = {r["data"]: r["valor"] for r in sgs["juros_sp_pct"]}
    for r in sgs["nfsp_sp"]:
        j = juros_sp_map.get(r["data"])
        if j is None or r["valor"] is None:
            continue
        primario_sp_pct.append({"data": r["data"], "valor_pct": round(j - r["valor"], 4)})

    # ── v2: r−g, estabilizador histórico e decomposição anual da DLSP ──
    print("  [v2] sustentabilidade (r implícita DLSP × g nominal) + decomposição da DLSP")
    sust_serie = calcula_sustentabilidade(sgs["juros_sp_pct"], sgs["dlsp_total"], primario_sp_pct, pib_map)
    decomposicao_dlsp = decompoe_dlsp_anual(sust_serie, sgs["dlsp_total"], sgs["juros_sp_pct"], primario_sp_pct)
    if sust_serie:
        u = sust_serie[-1]
        print(f"  [v2] {u['data']}: r={u['r_aa_pct']}% g={u['g_aa_pct']}% | r-g={u['r_menos_g_pp']}pp | p*={u['primario_estabilizador_pct_pib']}% PIB")
        assert -5 < u["r_aa_pct"] < 40, "taxa implícita fora de faixa plausível"
    if decomposicao_dlsp:
        res_max = max(abs(d["residuo_pp"]) for d in decomposicao_dlsp[-10:])
        print(f"  [v2] decomposição DLSP: {len(decomposicao_dlsp)} anos | resíduo máx (10a): {res_max:.2f}pp")

    payload = {
        "schema_version": 2,
        "gerado_em": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "mes_recente": sgs["dbgg"][-1]["data"] if sgs["dbgg"] else None,
        "pib_nominal_12m_brl_milhoes": pib_12m_recente,
        "divida": {
            "dbgg_pct_pib": sgs["dbgg"],
            "dlsp_total_pct_pib": sgs["dlsp_total"],
            "dlsp_gov_central_pct_pib": sgs["dlsp_gov_central"],
        },
        "receita_e_gastos": {
            "receita_liquida_12m_brl_mm": receita_liquida_12m,
            "despesa_total_12m_brl_mm": despesa_total_12m,
            "primario_central_12m_brl_mm": primario_central_12m,
            "juros_central_12m_brl_mm": juros_central_12m,
            "receita_liquida_pct_pib": receita_pct_pib,
            "despesa_total_pct_pib": despesa_pct_pib,
            "primario_central_pct_pib": primario_central_pct_pib,
            "juros_central_pct_pib": juros_central_pct_pib,
            "despesa_pct_receita": despesa_pct_rec,
            "juros_pct_receita": juros_pct_rec,
            "primario_pct_receita": primario_pct_rec,
            "previdencia_12m_pct_pib": previdencia_pct_pib,
            "pessoal_12m_pct_pib": pessoal_pct_pib,
            "previdencia_12m_pct_receita": previdencia_pct_rec,
            "pessoal_12m_pct_receita": pessoal_pct_rec,
            "discricionarias_12m_brl_mm": discricionarias_12m,
            "outras_obrigatorias_12m_brl_mm": outras_obrig_12m,
            "abono_seguro_12m_pct_pib": abono_seguro_pct_pib,
            "bpc_loas_12m_pct_pib": bpc_pct_pib,
            "fundeb_12m_pct_pib": fundeb_pct_pib,
            "subsidios_12m_pct_pib": subsidios_pct_pib,
            "discricionarias_12m_pct_pib": discricionarias_pct_pib,
            "outras_obrigatorias_12m_pct_pib": outras_obrig_pct_pib,
            "abono_seguro_12m_pct_receita": abono_seguro_pct_rec,
            "bpc_loas_12m_pct_receita": bpc_pct_rec,
            "fundeb_12m_pct_receita": fundeb_pct_rec,
            "subsidios_12m_pct_receita": subsidios_pct_rec,
            "discricionarias_12m_pct_receita": discricionarias_pct_rec,
            "outras_obrigatorias_12m_pct_receita": outras_obrig_pct_rec,
            # Decomposicao receita por tributo (% PIB)
            "imposto_renda_12m_pct_pib": ir_pct_pib,
            "cofins_12m_pct_pib": cofins_pct_pib,
            "csll_12m_pct_pib": csll_pct_pib,
            "pis_pasep_12m_pct_pib": pis_pct_pib,
            "ipi_12m_pct_pib": ipi_pct_pib,
            "iof_12m_pct_pib": iof_pct_pib,
            "imposto_importacao_12m_pct_pib": ii_pct_pib,
            "cide_12m_pct_pib": cide_pct_pib,
            "rgps_arrecadacao_12m_pct_pib": rgps_pct_pib,
            "concessoes_12m_pct_pib": concessoes_pct_pib,
            "dividendos_12m_pct_pib": dividendos_pct_pib,
            "recursos_naturais_12m_pct_pib": rec_naturais_pct_pib,
            "nfsp_sp_12m_pct_pib": sgs["nfsp_sp"],
            "primario_sp_12m_pct_pib": primario_sp_pct,
            "juros_nominais_sp_12m_pct_pib": sgs["juros_sp_pct"],
            "nominal_sp_12m_pct_pib": nominal_sp_pct,
        },
        "monetaria": {
            "selic_diaria_pct": selic_diaria[-730:],
            "ipca_12m_pct": sgs["ipca_12m"],
            "selic_real_ex_post_pct": selic_real,
            "pib_real_yoy_pct": pib_real_yoy_serie,
        },
        "composicao_dpmfi": {
            "selic_pct": sgs["dpmfi_pct_selic"],
            "prefixado_pct": sgs["dpmfi_pct_prefixado"],
            "cambio_pct": sgs["dpmfi_pct_cambio"],
            "tr_pct": sgs["dpmfi_pct_tr"],
            "outros_pct": sgs["dpmfi_pct_outros"],
            # v2: a fatia que faltava (≈2ª maior) — SGS 12001; o stack fecha em ~100%
            "indices_precos_pct": sgs["dpmfi_pct_indices_precos"],
        },
        "credito_economia": {
            "credito_total_pct_pib": sgs["credito_total_pct_pib"],
        },
        "stress": {
            "reer_index": sgs["reer"],
            "reservas_usd_mm_mensal": reservas_mensal,
        },
        "pib": {
            "acumulado_12m_brl_milhoes_mensal": sgs["pib_12m_brl"],
            "real_idx": sgs["pib_real_idx"],
        },
        "expectativas_focus": {
            "selic_anuais": {str(k): v for k, v in focus_selic.items()},
            "ipca_anuais": {str(k): v for k, v in focus_ipca.items()},
            "pib_anuais": {str(k): v for k, v in focus_pib.items()},
            "cambio_anuais": {str(k): v for k, v in focus_cambio.items()},
        },
        "metas_ldo": {
            "_fonte": "Metas de primário do GOVERNO CENTRAL por LDO. Trajetória vigente desde o PLDO 2025: 2024 = 0,00 / 2025 = 0,00 / 2026 = +0,25 / 2027 = +0,50 (% PIB), banda ±0,25pp (LC 200/2023; banda só existe a partir de 2024). A meta de 2023 era fixada em R$ (LDO 2023 alterada), sem banda do arcabouço — por isso não entra aqui. AUDITAR ANUALMENTE contra a LDO do ano. Convenção: positivo = superávit. A aferição oficial é no ANO-CALENDÁRIO com abatimentos (ex.: precatórios EC 114) — comparar com 12m móvel é aproximação.",
            "anos": {
                "2024": {"centro": 0.00, "banda_inf": -0.25, "banda_sup": 0.25},
                "2025": {"centro": 0.00, "banda_inf": -0.25, "banda_sup": 0.25},
                "2026": {"centro": 0.25, "banda_inf": 0.00, "banda_sup": 0.50},
                "2027": {"centro": 0.50, "banda_inf": 0.25, "banda_sup": 0.75}
            }
        },
        # ── v2 ──
        "sustentabilidade": {
            "_perimetro": "Setor público consolidado: r = taxa implícita da DLSP (juros nominais 12m ÷ DLSP média 12m, ambos SGS); g = PIB nominal acumulado 12m YoY (SGS 4382); p* = (r−g)/(1+g) × DLSP%PIB t-12. Perímetro ÚNICO — juros consolidados não acruam sobre a DBGG (governo geral). Fórmula calculada SÓ aqui; front não recalcula.",
            "serie": sust_serie,
        },
        "decomposicao_dlsp": {
            "_nota": "Δ DLSP anual ≈ juros − primário − efeito crescimento + resíduo (ajustes patrimoniais/cambiais). Identidade limpa no consolidado; a decomposição OFICIAL da DBGG (fatores condicionantes, Nota de Imprensa BCB) será adicionada quando coletada.",
            "anos": decomposicao_dlsp,
        },
        "receita_familias": {
            "_nota": "Agregados de família do próprio RTN (linhas 1.1/1.2/1.3/1.4) — o stack fecha com a receita TOTAL; receita líquida = total − transferências a E&M.",
            "administrada_rfb_12m_pct_pib": admin_pct_pib,
            "incentivos_fiscais_12m_pct_pib": incentivos_pct_pib,
            "rgps_12m_pct_pib": rgps_pct_pib,
            "nao_administrada_12m_pct_pib": nao_admin_pct_pib,
            "dividendos_concessoes_12m_pct_pib": dividendos_concessoes_pct_pib,
        },
        "despesa_rubricas_v2": {
            "_nota": "Residual 'demais obrigatórias' = 4.3 − (abono + BPC + FUNDEB + subsídios): somar o 4.3 inteiro duplicaria as sub-rubricas já plotadas. Com previdência, pessoal, 4 sub-rubricas, residual, obrigatórias c/ controle de fluxo (4.4.1) e discricionárias (4.4.2), o stack fecha com a despesa total.",
            "demais_obrigatorias_12m_pct_pib": demais_obrig_pct_pib,
            "obrig_controle_fluxo_12m_pct_pib": obrig_fluxo_pct_pib,
        },
        "arcabouco": {
            "_nota": "Limite legal (LC 200/2023): crescimento real da despesa do ano t = 70% do crescimento real da receita 12m até jun/t-1, entre 0,6% e 2,5% a.a. (apurado com IPCA). O corredor 0,6–2,5 é plotável; o limite EXATO de cada ano sai na LDO/Relatório bimestral — conferir antes de afirmar violação.",
            "corredor": {"piso_pct": 0.6, "teto_pct": 2.5},
            "despesa_real_12m_yoy_pct": despesa_real_yoy,
            "receita_real_12m_yoy_pct": receita_real_yoy,
        },
        "acompanhamento_meta": {
            "_nota": "Primário do governo central acumulado jan→mês (acima da linha), como o RTN publica — a meta LDO é aferida no ano-calendário, não em 12m móvel.",
            "primario_central_ytd_brl_mm": primario_ytd,
        },
        "destaques": {
            "dpmfi_selic_pct_recente": last_val(sgs["dpmfi_pct_selic"]),
            "dpmfi_prefixado_pct_recente": last_val(sgs["dpmfi_pct_prefixado"]),
            "dpmfi_cambio_pct_recente": last_val(sgs["dpmfi_pct_cambio"]),
            "credito_total_pct_pib_recente": last_val(sgs["credito_total_pct_pib"]),
            "dbgg_pct_recente": last_val(sgs["dbgg"]),
            "dlsp_pct_recente": last_val(sgs["dlsp_total"]),
            "receita_liquida_pct_pib_recente": last_val(receita_pct_pib, "valor_pct"),
            "despesa_total_pct_pib_recente": last_val(despesa_pct_pib, "valor_pct"),
            "primario_central_pct_pib_recente": last_val(primario_central_pct_pib, "valor_pct"),
            "juros_central_pct_pib_recente": last_val(juros_central_pct_pib, "valor_pct"),
            "juros_pct_receita_recente": last_val(juros_pct_rec, "valor_pct"),
            "primario_pct_receita_recente": last_val(primario_pct_rec, "valor_pct"),
            "nfsp_sp_pct_recente": last_val(sgs["nfsp_sp"]),
            "reer_recente": last_val(sgs["reer"]),
            "reservas_usd_recente": last_val(reservas_mensal),
            "selic_real_recente": last_val(selic_real, "selic_real_pct"),
            "pib_real_yoy_recente": last_val(pib_real_yoy_serie, "valor_yoy_pct"),
            "ipca_12m_recente": last_val(sgs["ipca_12m"]),
        },
    }

    out_file.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    size = out_file.stat().st_size
    print(f"\n  -> {out_file} ({size / 1024:.1f} KB)")

    # Series essenciais: um dia ruim do api.bcb.gov.br nao pode zerar DBGG/DLSP
    # no ar. Se alguma estiver vazia, recusa o upload (exit != 0; o retry diario
    # do workflow cuida do resto) e o dado bom continua no Blob.
    essenciais = {"dbgg": sgs["dbgg"], "dlsp_total": sgs["dlsp_total"]}
    vazias = [nome for nome, serie in essenciais.items() if not serie]
    if vazias:
        print(
            f"[ERROR] series essenciais vazias ({', '.join(vazias)}) — upload abortado "
            f"para preservar o dado existente no Blob",
            file=sys.stderr,
        )
        sys.exit(1)

    if args.upload:
        maybe_upload_json(out_file, BLOB_PATH)


if __name__ == "__main__":
    main()
