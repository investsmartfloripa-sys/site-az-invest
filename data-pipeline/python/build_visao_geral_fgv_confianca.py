"""Build do JSON do Painel Visao Geral - bloco Confianca FGV-IBRE.

Indices de Confianca da FGV-IBRE:
- ICE - Confianca Empresarial composto
- ICI - Industria
- ICOM - Comercio
- ICS - Servicos
- ICST - Construcao
- ICA - Agropecuaria
- ICC - Consumidor

Cada um normalmente vive em pagina propria do portal IBRE com link XLSX.
Padrao defensivo: tenta cada pagina, parseia o XLSX, sao todos opcionais.

NOTA: assinatura FGVDados tem mais granularidade. Esse pipeline pega
o headline mensal disponivel publicamente.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests

HERE = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = (HERE.parent / "out").resolve()
BLOB_PATH = "data/visao_geral_fgv_confianca.json"
UA = {"User-Agent": "az-invest-visao-geral-fgv-conf/0.1"}

# URLs publicas (sujeitas a mudancas - parser defensivo)
PAGES = {
    "ice": "https://portalibre.fgv.br/en/confianca-empresarial",
    "ici": "https://portalibre.fgv.br/en/sondagem-industria",
    "icom": "https://portalibre.fgv.br/en/sondagem-comercio",
    "ics": "https://portalibre.fgv.br/en/sondagem-servicos",
    "icst": "https://portalibre.fgv.br/en/sondagem-construcao",
    "ica": "https://portalibre.fgv.br/en/sondagem-agropecuaria",
    "icc": "https://portalibre.fgv.br/en/sondagem-consumidor",
}

INPUTS = {
    "ice": "2005-09",
    "ici": "1995-12",
    "icom": "2005-09",
    "ics": "2008-06",
    "icst": "2010-06",
    "ica": "2014-01",
    "icc": "2005-09",
}


def _get(url: str, *, timeout: int = 60, retries: int = 2) -> requests.Response | None:
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=UA)
            r.raise_for_status()
            return r
        except Exception:
            time.sleep(3)
    return None


def localizar_xlsx(page_url: str, palavras: list[str]) -> str | None:
    r = _get(page_url)
    if r is None:
        return None
    urls = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', r.text)
    if not urls:
        return None
    for u in urls:
        if any(p in u.lower() for p in palavras):
            if u.startswith("//"):
                u = "https:" + u
            elif u.startswith("/"):
                u = "https://portalibre.fgv.br" + u
            return u
    u = urls[0]
    if u.startswith("//"):
        u = "https:" + u
    elif u.startswith("/"):
        u = "https://portalibre.fgv.br" + u
    return u


MES_PT = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


def parse_data(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.match(r"^([A-Za-z]{3})[/\-\s](\d{2,4})$", s)
    if m:
        mes = MES_PT.get(m.group(1).upper())
        if mes:
            ano = int(m.group(2))
            ano = ano + 2000 if ano < 100 else ano
            return f"{ano:04d}-{mes:02d}"
    return None


def parse_xlsx_indicador(content: bytes, palavra_chave: str) -> dict[str, float]:
    """Tenta achar a serie do indicador no XLSX. Aceita 'sondagem', 'confianca', etc."""
    from openpyxl import load_workbook

    out: dict[str, float] = {}
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return out
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        # acha header e coluna de "indice" ou nome do indicador
        header_idx = None
        for i, row in enumerate(rows[:25]):
            row_str = [str(c).upper() if c is not None else "" for c in row]
            joined = " ".join(row_str)
            if palavra_chave.upper() in joined or "INDICE" in joined or "INDEX" in joined:
                header_idx = i
                break
        if header_idx is None:
            continue
        header = [str(c).strip().upper() if c is not None else "" for c in rows[header_idx]]
        col_data = 0
        col_val = None
        for i, c in enumerate(header):
            if palavra_chave.upper() in c and i != 0:
                col_val = i
                break
        if col_val is None:
            # cair pra primeira coluna numerica nao-data
            col_val = 1 if len(header) > 1 else None
        if col_val is None:
            continue
        for row in rows[header_idx + 1 :]:
            if len(row) <= col_val:
                continue
            mes = parse_data(row[col_data])
            if not mes:
                continue
            try:
                out[mes] = float(row[col_val])
            except (TypeError, ValueError):
                continue
        if out:
            break
    return out


def serie_lista(d: dict[str, float]) -> list[dict]:
    return [{"mes": m, "valor": d[m]} for m in sorted(d.keys())]


def buscar_indicador(slug: str, palavra: str) -> dict[str, float]:
    print(f"  -> {slug.upper()}")
    page = PAGES.get(slug)
    if not page:
        return {}
    url = localizar_xlsx(page, [slug, palavra, "sondagem"])
    if not url:
        print(f"    XLSX nao localizado em {page}")
        return {}
    print(f"    XLSX: {url}")
    r = _get(url)
    if r is None:
        return {}
    return parse_xlsx_indicador(r.content, palavra)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--soft-fail", action="store_true")
    args = ap.parse_args()

    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "visao_geral_fgv_confianca.json"

    print("== FGV-IBRE - Confiancas ==")
    series: dict[str, dict[str, float]] = {}
    series["ice"] = buscar_indicador("ice", "confianca")
    series["ici"] = buscar_indicador("ici", "industria")
    series["icom"] = buscar_indicador("icom", "comercio")
    series["ics"] = buscar_indicador("ics", "servicos")
    series["icst"] = buscar_indicador("icst", "construcao")
    series["ica"] = buscar_indicador("ica", "agropecuaria")
    series["icc"] = buscar_indicador("icc", "consumidor")

    any_serie = any(series.values())
    if not any_serie:
        print("  Nenhuma serie obtida — preservando Blob", file=sys.stderr)
        sys.path.insert(0, str(HERE))
        from shared.blob_download import download_json
        prev = download_json(BLOB_PATH)
        if prev:
            prev["freshness_status"] = "stale"
            prev["gerado_em"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            out_file.write_text(json.dumps(prev, indent=2, ensure_ascii=False), encoding="utf-8")
            return
        payload = {
            "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "freshness_status": "missing",
            **{k: [] for k in series},
            "metadata": {"fonte": "FGV-IBRE", "nota": "Scraper nao localizou XLSX. Onda 2 sob revisao manual."},
        }
        out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return

    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "freshness_status": "fresh",
        "ice": serie_lista(series["ice"]),
        "ici": serie_lista(series["ici"]),
        "icom": serie_lista(series["icom"]),
        "ics": serie_lista(series["ics"]),
        "icst": serie_lista(series["icst"]),
        "ica": serie_lista(series["ica"]),
        "icc": serie_lista(series["icc"]),
        "inputs": INPUTS,
        "min_start_date": max(INPUTS.values()),
        "metadata": {
            "fonte": "FGV-IBRE - Sondagens. Headline mensal publico do portal IBRE.",
            "nota": "100 = neutro. >100 otimismo, <100 pessimismo. Estrutura XLSX pode mudar - parser defensivo.",
        },
    }
    out_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON {out_file} ({out_file.stat().st_size / 1024:.1f} KB)")
    for k, s in series.items():
        print(f"  {k.upper()} {len(s)} obs")

    if args.upload:
        sys.path.insert(0, str(HERE))
        from shared.blob_upload import maybe_upload_json
        try:
            maybe_upload_json(out_file, BLOB_PATH)
        except Exception as e:
            print(f"[upload] FALHOU: {e}", file=sys.stderr)
            if not args.soft_fail:
                sys.exit(1)


if __name__ == "__main__":
    main()
